"""Convert PROJECT LIST.xlsm into a JS file the portal can load directly.

The portal's Project List view aggregates bids where outcome == "Awarded" by
project name, so each row from the workbook becomes a single bid record with
outcome=Awarded / status=Archived. The portal then groups, numbers and totals
them itself.

Output is `window.getStaticProjectBids = function(){ return [...]; };` so the
file can be loaded with a normal <script> tag and the data is available on the
window before app startup runs.
"""
import argparse
import json
import re
import warnings
from datetime import datetime
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

# Header row in PROJECT LIST sheet (1-based)
HEADER_ROW = 3
DATA_START_ROW = 4
SHEET_NAME = "PROJECT LIST"

# Column letter -> portal field. Uses the headers I confirmed from the workbook.
COL_MAP = {
    "B": "fullNumber",        # FULL # FOR FD (e.g. "2611-BAY")
    "C": "projectName",       # PROJECT NAME
    "D": "estNumber",         # JOB # (we treat the job # as the est # in the portal)
    "E": "link",              # LINK
    "F": "pmOpenTasks",       # PM OPEN TASKS
    "G": "peOpenTasks",       # PE OPEN TASKS
    "H": "admOpenTasks",      # ADM / ACCT OPEN TASKS
    "I": "projectManager",    # PROJECT MANAGER
    "J": "projectEngineer",   # PROJECT ENGINEER
    "K": "generalForeman",    # GENERAL FOREMAN / SUPER
    "L": "projectForeman",    # PROJECT FOREMAN
    "M": "clientGc",          # CUSTOMER
    "N": "division",          # S/B (BAY / SAC)
    "O": "cprDir",            # CPR / DIR #
    "P": "projectType",       # JOB TYPE
    "Q": "local",             # LOCAL (union local #)
    "R": "foundationStatus",  # FOUNDATION STATUS
    "S": "awardDate",         # AWARD DATE
    "T": "startDate",         # START DATE
    "U": "duration",          # DURATION
    "V": "endDate",           # END DATE
    "W": "contactOffice",     # CUSTOMER CONTACT OFFICE
    "X": "contactField",      # CUSTOMER CONTACT FIELD
    "Y": "originalEstNumber", # EST # (the original estimate # before award)
    "Z": "originalContract",  # ORIGINAL CONTRACT
    "AH": "jobListStatus",    # JOB LIST STATUS
    "BE": "jobAddress",       # JOB ADDRESS
}

# Sentinels for empty data
EMPTY_VALUES = {"", "0", "00:00:00", "n/a", "tbd", "no"}


def col_letter_to_index(letter):
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch.upper()) - ord("A") + 1)
    return n - 1  # 0-based for tuple indexing


def read_cell(row_tuple, col_letter):
    idx = col_letter_to_index(col_letter)
    if idx >= len(row_tuple):
        return None
    return row_tuple[idx]


def to_iso_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        # Excel dates often have a 00:00:00 time; treat as the date portion
        return value.date().isoformat()
    s = str(value).strip()
    if not s or s.lower() in EMPTY_VALUES:
        return None
    # Some rows store dates as strings like "2026-01-13 00:00:00"
    try:
        return datetime.fromisoformat(s.split(" ")[0]).date().isoformat()
    except Exception:
        return None


def to_int(value):
    if value is None:
        return 0
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        try:
            return int(value)
        except Exception:
            return 0
    s = re.sub(r"[^0-9\-]", "", str(value))
    try:
        return int(s) if s else 0
    except Exception:
        return 0


def to_money(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(s) if s else 0.0
    except Exception:
        return 0.0


def clean_str(value):
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in EMPTY_VALUES:
        return ""
    return s


def normalize_division(value):
    s = clean_str(value).upper()
    if s.startswith("BAY"):
        return "BAY"
    if s.startswith("SAC"):
        return "SAC"
    return s


def stable_id(prefix, *parts):
    raw = "|".join([str(p or "") for p in parts])
    return f"{prefix}-{re.sub(r'[^a-zA-Z0-9]+', '-', raw).strip('-').lower()}"


LINK_COL_LETTER = "E"

# Each individual per-job sheet stores Change Orders starting at row 18 in
# columns K..R, with headers on row 17:
#   K=CO# | L=DESCRIPTION | M=APPROVAL BY | N=FD & FM SCOPE
#   O=BUDGET HRS | P=PRICE | Q=STATUS | R=Due
CO_HEADER_ROW = 17
CO_COL_LETTERS = {
    "number": "K",
    "title": "L",
    "approval": "M",
    "scope": "N",
    "hours": "O",
    "amount": "P",
    "status": "Q",
}
CO_MAX_ROWS = 30  # CO grid is bounded; later rows contain Submittal/RFI/etc. sections

# When walking down the CO section, these patterns in the K (CO#) or L
# (DESCRIPTION) columns mark the start of the *next* section in the per-job
# sheet — Submittals, RFIs, Tasks, Material Account, etc. Stop here.
CO_STOP_K_PATTERNS = [
    "SM STATUS",
    "MATERIAL ACCOUNT",
    "SUB TOTAL",
    "SUBS / PO",
    "SUBS/PO",
    "SCHEDULE TRACKING",
    "PHASE #",
    "PHASE NAME",
]
CO_STOP_K_EXACT = {"RFI", "TASK"}
CO_STOP_L_PATTERNS = [
    "SUBMITTAL",
    "NAME / DESCRIPTION",
    "NAME/DESCRIPTION",
    "SCOPE DESCRIPTION",
    "PHASE START DATE",
]


def is_co_stop_row(k_value, l_value):
    k_norm = clean_str(k_value).upper()
    l_norm = clean_str(l_value).upper()
    if k_norm in CO_STOP_K_EXACT:
        return True
    if any(p in k_norm for p in CO_STOP_K_PATTERNS):
        return True
    if any(p in l_norm for p in CO_STOP_L_PATTERNS):
        return True
    return False


def portal_slug(value):
    """Mirror the portal's slugify(): lowercase, non-alnum -> '-', trim, slice 32."""
    s = str(value or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = s.strip("-")
    return s[:32]


def normalize_co_status(value):
    s = clean_str(value).lower()
    if not s:
        return "Pending"
    if "approved" in s and "pending" not in s:
        return "Approved"
    if "rejected" in s or "denied" in s:
        return "Rejected"
    if "approval pending" in s or "pending" in s:
        return "Pending"
    return clean_str(value).title()


def extract_change_orders(wb, job_number):
    """Return a list of CO dicts the portal can drop straight into projectCoStore."""
    if not job_number:
        return []
    if job_number not in wb.sheetnames:
        return []
    ws = wb[job_number]

    cols = {key: col_letter_to_index(letter) + 1 for key, letter in CO_COL_LETTERS.items()}

    out = []
    seen_ids = set()
    end_row = min(CO_HEADER_ROW + CO_MAX_ROWS, ws.max_row)
    for row_idx in range(CO_HEADER_ROW + 1, end_row + 1):
        k_val = ws.cell(row_idx, cols["number"]).value
        l_val = ws.cell(row_idx, cols["title"]).value
        if is_co_stop_row(k_val, l_val):
            break
        number = clean_str(k_val)
        title = clean_str(l_val)
        if not number and not title:
            continue
        amount = to_money(ws.cell(row_idx, cols["amount"]).value)
        status = normalize_co_status(ws.cell(row_idx, cols["status"]).value)
        approval = clean_str(ws.cell(row_idx, cols["approval"]).value)
        scope = clean_str(ws.cell(row_idx, cols["scope"]).value)
        hours = to_money(ws.cell(row_idx, cols["hours"]).value)
        display_title = title or f"CO {number}"
        if number:
            display_title = f"#{number} — {display_title}" if title else f"CO #{number}"
        co_id = stable_id("co", job_number, number, title) or f"co-{job_number}-{row_idx}"
        if co_id in seen_ids:
            co_id = f"{co_id}-{row_idx}"
        seen_ids.add(co_id)
        iso_now = datetime.utcnow().isoformat() + "Z"
        out.append({
            "id": co_id,
            "title": display_title,
            "amount": amount,
            "status": status,
            "approval": approval,
            "scope": scope,
            "budgetHours": hours,
            "createdAt": iso_now,
            "updatedAt": iso_now,
        })
    return out


def row_to_record(row_tuple, link_url=None):
    rec = {key: read_cell(row_tuple, col) for col, key in COL_MAP.items()}
    project_name = clean_str(rec["projectName"])
    job_number = clean_str(rec["estNumber"])
    if not project_name or not job_number:
        return None

    award_date = to_iso_date(rec["awardDate"])
    start_date = to_iso_date(rec["startDate"])
    end_date = to_iso_date(rec["endDate"])
    contract = to_money(rec["originalContract"])
    division = normalize_division(rec["division"])

    contact_parts = [clean_str(rec["contactOffice"]), clean_str(rec["contactField"])]
    contact = " / ".join([p for p in contact_parts if p])

    pm = clean_str(rec["projectManager"])
    pe = clean_str(rec["projectEngineer"])
    general_foreman = clean_str(rec["generalForeman"])
    project_foreman = clean_str(rec["projectForeman"])
    job_list_status = clean_str(rec["jobListStatus"])

    link_raw = clean_str(rec["link"])
    # Excel commonly stores "LINK" as a hyperlink with display text "LINK". The
    # iter_rows values_only=True path only gives us the display string, so the
    # caller passes the resolved hyperlink target separately.
    link_display = link_raw or ("LINK" if link_url else "")

    iso_now = datetime.utcnow().isoformat() + "Z"
    bid = {
        "id": stable_id("proj", job_number, project_name),
        "estNumber": job_number,
        "originalEstNumber": clean_str(rec["originalEstNumber"]),
        "fullNumber": clean_str(rec["fullNumber"]),
        "projectName": project_name,
        "clientGc": clean_str(rec["clientGc"]),
        "contact": contact,
        "contactOffice": clean_str(rec["contactOffice"]),
        "contactField": clean_str(rec["contactField"]),
        "location": clean_str(rec["jobAddress"]),
        "scope": "",
        "estimator": pm or pe,
        "projectManager": pm,
        "projectEngineer": pe,
        "generalForeman": general_foreman,
        "projectForeman": project_foreman,
        "division": division,
        "projectType": clean_str(rec["projectType"]),
        "local": clean_str(rec["local"]),
        "cprDir": clean_str(rec["cprDir"]),
        "foundationStatus": clean_str(rec["foundationStatus"]),
        "jobListStatus": job_list_status,
        "duration": clean_str(rec["duration"]),
        "bidDueDate": start_date or award_date,
        "jobWalk": "",
        "bidAmount": contract,
        "awardedAmount": contract,
        "status": "Archived",
        "outcome": "Awarded",
        "notes": "",
        "dateSent": start_date or award_date,
        "followUpDate": None,
        "dateAwarded": award_date,
        "startDate": start_date,
        "endDate": end_date,
        "linkLabel": link_display,
        "linkUrl": link_url or "",
        "openTasks": {
            "pm": to_int(rec["pmOpenTasks"]),
            "pe": to_int(rec["peOpenTasks"]),
            "adm": to_int(rec["admOpenTasks"]),
        },
        "createdAt": iso_now,
        "updatedAt": iso_now,
    }
    return bid


def parse_workbook(wb, *, source_label="(unknown)"):
    """Parse a loaded PROJECT LIST.xlsm workbook into the portal payload dict.

    Extracted so both the local pipeline (loads from a path) and the
    cloud-side pipeline (loads from SharePoint share-link bytes) share
    one parse implementation. Per-project change orders come from a
    separate worksheet named after each estNumber, all walked here.
    """
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")
    ws = wb[SHEET_NAME]

    link_col_idx_excel = col_letter_to_index(LINK_COL_LETTER) + 1  # openpyxl is 1-based

    bids = []
    seen_ids = set()
    change_orders = {}  # portal-project-key -> [CO dicts]
    co_count = 0
    for row_cells in ws.iter_rows(min_row=DATA_START_ROW):
        row_tuple = tuple(c.value for c in row_cells)
        link_url = None
        # openpyxl gives us a sparse row; index defensively
        for c in row_cells:
            if c.column == link_col_idx_excel and c.hyperlink and c.hyperlink.target:
                link_url = c.hyperlink.target
                break
        rec = row_to_record(row_tuple, link_url=link_url)
        if not rec:
            continue
        if rec["id"] in seen_ids:
            continue
        seen_ids.add(rec["id"])
        bids.append(rec)

        # Pull change orders from the per-job sheet keyed by JOB#.
        cos = extract_change_orders(wb, rec["estNumber"])
        if cos:
            # Use the same slug the portal computes for project.key (slugify of
            # project name). This is what projectCoStore[key] looks up.
            project_key = portal_slug(rec["projectName"]) or portal_slug(rec["estNumber"]) or rec["id"]
            change_orders[project_key] = cos
            co_count += len(cos)

    return {
        "generatedAt": datetime.utcnow().isoformat() + "Z",
        "source": source_label,
        "count": len(bids),
        "coCount": co_count,
        "bids": bids,
        "changeOrders": change_orders,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.excel, read_only=False, data_only=True)
    payload = parse_workbook(wb, source_label=str(args.excel))

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload_json = json.dumps(payload, indent=2)
    js = (
        "// AUTO-GENERATED — do not edit. Regenerate with fusion-portal/Convert-ProjectList.ps1\n"
        f"window.__STATIC_PROJECTS__ = {payload_json};\n"
        "window.getStaticProjectBids = function(){\n"
        "  var p = window.__STATIC_PROJECTS__;\n"
        "  return (p && Array.isArray(p.bids)) ? p.bids : [];\n"
        "};\n"
    )
    out_path.write_text(js, encoding="utf-8")
    # Also write a JSON sibling so the running page can poll for updates
    # without re-evaluating the JS file.
    json_path = out_path.with_suffix(".json")
    json_path.write_text(payload_json, encoding="utf-8")
    print(
        f"Wrote {len(payload['bids'])} project bids and {payload['coCount']} change orders across "
        f"{len(payload['changeOrders'])} projects to {out_path} and {json_path}"
    )


if __name__ == "__main__":
    main()
