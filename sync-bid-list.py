import argparse
import json
import re
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import openpyxl

DEFAULT_DIVISION = "ALL"
# Fallback mapping only if no division info exists for an estimator
BAY_ESTIMATORS = {"ALEXANDER TOLER", "DAVID DUENKEL", "JACOB DUENKEL", "JADE SUEKI", "GABRIEL THOMAS"}
# Hard overrides for known misclassified estimators
DIVISION_OVERRIDES = {"AUSTIN CARMICHAEL": "SAC"}

def load_core_estimators(html_path):
    try:
        text = Path(html_path).read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return None
    m = re.search(r"CORE_ESTIMATORS\s*=\s*\[([^\]]*)\]", text)
    if not m:
        return None
    block = m.group(1)
    names = re.findall(r"['\"]([^'\"]+)['\"]", block)
    return [n.strip() for n in names if n.strip()]


def update_core_estimators(html_path, names):
    text = Path(html_path).read_text(encoding="utf-8", errors="ignore")
    pattern = r"CORE_ESTIMATORS\s*=\s*\[[^\]]*\]"
    if not re.search(pattern, text):
        raise SystemExit("CORE_ESTIMATORS not found in HTML")
    sorted_names = sorted({n.strip() for n in names if n and str(n).strip()}, key=lambda s: s.lower())
    replacement = "CORE_ESTIMATORS = [" + ",".join([f'\"{n}\"' for n in sorted_names]) + "]"
    text = re.sub(pattern, replacement, text, count=1)
    Path(html_path).write_text(text, encoding="utf-8")
    return sorted_names


def estimator_allowed(name, allowed):
    if not allowed:
        return True
    n = str(name or "").strip().lower()
    return any(n == a.lower() for a in allowed)


def excel_date_to_iso(value):
    if value is None:
        return None
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    try:
        return pd.to_datetime(value).date().isoformat()
    except Exception:
        return None


def safe_number(value):
    if value is None or pd.isna(value):
        return 0
    try:
        return float(re.sub(r"[^0-9.\-]", "", str(value)))
    except Exception:
        return 0


def coerce_bool(value):
    if value is None or pd.isna(value):
        return False
    if isinstance(value, (int, float)):
        return value > 0
    s = str(value).strip().lower()
    if not s or s in ("0", "no", "false", "n"):
        return False
    return True


def normalize_division(value):
    s = str(value or "").strip().upper()
    if not s:
        return ""
    if s.startswith("BAY"):
        return "BAY"
    if s.startswith("SAC"):
        return "SAC"
    return s


def division_from_estimator(estimator):
    name = str(estimator or "").strip().upper()
    if not name:
        return ""
    if name in DIVISION_OVERRIDES:
        return DIVISION_OVERRIDES[name]
    return "BAY" if name in BAY_ESTIMATORS else "SAC"


def build_estimator_division_map(dfs_with_weights):
    counts = {}
    for df, weight in dfs_with_weights:
        if df is None or df.empty:
            continue
        if "DIVISION" not in df.columns:
            continue
        est_col = "ESTIMATOR" if "ESTIMATOR" in df.columns else ("Estimator" if "Estimator" in df.columns else None)
        if not est_col:
            continue
        for _, row in df.iterrows():
            est = str(row.get(est_col, "") or "").strip().upper()
            div = normalize_division(row.get("DIVISION", ""))
            if not est or not div:
                continue
            counts.setdefault(est, {})
            counts[est][div] = counts[est].get(div, 0) + weight
    mapping = {}
    for est, div_counts in counts.items():
        # pick highest weighted count
        best_div = None
        best_count = -1
        for div, c in div_counts.items():
            if c > best_count:
                best_div = div
                best_count = c
        if best_div:
            mapping[est] = best_div
    # apply overrides
    for est, div in DIVISION_OVERRIDES.items():
        mapping[est] = div
    return mapping


def row_value(row, keys):
    for key in keys:
        if key in row and not pd.isna(row[key]) and str(row[key]).strip() != "":
            return row[key]
    return None


def add_days_iso(iso_date, days):
    if not iso_date:
        return None
    try:
        d = datetime.fromisoformat(iso_date)
    except Exception:
        return None
    return (d + timedelta(days=days)).date().isoformat()


def outcome_from_status(value):
    s = str(value or "").lower()
    if "award" in s and "not" not in s:
        return "Awarded"
    if "not" in s:
        return "Not Awarded"
    return "Pending"


def _build_documents_url_map(excel_path):
    """Read column B hyperlinks across BIDS / FOLLOW UPS / ARCHIVE sheets,
    keyed by est_number (uppercase). pandas strips Excel hyperlink data so
    we have to re-open with openpyxl to grab cell.hyperlink.target.
    Each sheet has its own header row offset (BIDS=8, FOLLOW UPS=4,
    ARCHIVE=2) — we look at the row immediately after the header for the
    est-number column letter and assume a stable column layout."""
    out = {}
    sheet_configs = [
        ("BIDS",        8,  "B", "C"),
        ("FOLLOW UPS",  4,  "B", "C"),
        ("ARCHIVE",     2,  "B", "C"),
    ]
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as exc:
        print(f"[warn] Could not open workbook for hyperlink scan: {exc}")
        return out
    for sheet_name, data_start, name_col, est_col in sheet_configs:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in range(data_start, ws.max_row + 1):
            name_cell = ws[f"{name_col}{row}"]
            est_cell  = ws[f"{est_col}{row}"]
            if not name_cell.hyperlink:
                continue
            est = str(est_cell.value or "").strip().upper()
            if not est:
                continue
            url = name_cell.hyperlink.target
            if url and est not in out:
                out[est] = url
    wb.close()
    return out


def parse_bids(df, allowed, pe_map=None):
    results = []
    for _, row in df.iterrows():
        estimator = str(row_value(row, ["ESTIMATOR", "Estimator"]) or "").strip()
        if not estimator or not estimator_allowed(estimator, allowed):
            continue
        project = row_value(row, ["PROJECT/ESTIMATE NAME", "PROJECT/ESTIMATE  NAME", "Project"])
        est_number = row_value(row, ["EST #", "EST#", "Est #"])
        bid_date = excel_date_to_iso(row_value(row, ["BID DATE", "Bid Date"]))
        if not project or not est_number or not bid_date:
            continue
        status_raw = str(row_value(row, ["STATUS", "Status"]) or "Bidding").strip()
        bid_amount = safe_number(row_value(row, ["WOULD HAVE WON BID AT", "Bid Amount", "BID AMOUNT"]))
        project_engineer = str(row_value(row, ["PE", "Project Engineer", "Project Eng"]) or "").strip()
        if not project_engineer and pe_map:
            project_engineer = pe_map.get(estimator, "")
        division = normalize_division(row_value(row, ["DIVISION", "Division"])) or division_from_estimator(estimator)
        results.append({
            "estNumber": str(est_number).strip(),
            "projectName": str(project).strip(),
            "clientGc": str(row_value(row, ["CLIENT", "Client"]) or "").strip(),
            "contact": "",
            "location": "",
            "scope": "",
            "jobWalk": str(row_value(row, ["JOB WALK", "Job Walk"]) or "").strip(),
            "estimator": estimator,
            "projectEngineer": project_engineer,
            "division": division,
            "bidDueDate": bid_date,
            "bidAmount": bid_amount,
            "awardedAmount": 0,
            "itemDueMod": safe_number(row_value(row, ["ITEM DUE MOD", "ITEM DUE MOD."])),
            "bidDueTime": str(row_value(row, ["TIME DUE", "BID TIME", "DUE TIME", "Time Due"]) or "").strip(),
            "setup": coerce_bool(row_value(row, ["SETUP"])),
            "documentsReview": coerce_bool(row_value(row, ["DOCUMENTS REVIEW"])),
            "requestQuotes": coerce_bool(row_value(row, ["REQUEST QUOTES"])),
            "takeoffFunding": coerce_bool(row_value(row, ["TAKEOFF & FUNDING", "TAKEOFF AND FUNDING"])),
            "addendaReceived": safe_number(row_value(row, ["ADDENDA RECEIVED"])),
            "addendaConfirmed": safe_number(row_value(row, ["ADDENDA CONFIRMED"])),
            "finalizeNumbers": coerce_bool(row_value(row, ["FINALIZE NUMBERS"])),
            "finalBidCheck": coerce_bool(row_value(row, ["FINAL BID CHECK"])),
            "status": status_raw or "Bidding",
            "outcome": "Pending",
            "notes": "",
            "dateSent": None,
            "followUpDate": None,
            "dateAwarded": None,
        })
    return results


def parse_follow(df, allowed, pe_map=None):
    results = []
    for _, row in df.iterrows():
        estimator = str(row_value(row, ["ESTIMATOR", "Estimator"]) or "").strip()
        if not estimator or not estimator_allowed(estimator, allowed):
            continue
        project = row_value(row, ["PROJECT/ESTIMATE NAME", "PROJECT/ESTIMATE  NAME", "Project"])
        est_number = row_value(row, ["EST #", "EST#", "Est #"])
        bid_date = excel_date_to_iso(row_value(row, ["BID DATE", "Bid Date"]))
        if not project or not est_number or not bid_date:
            continue
        bid_amount = safe_number(row_value(row, ["PRICE", "Bid Amount", "BID AMOUNT"]))
        project_engineer = str(row_value(row, ["PE", "Project Engineer", "Project Eng"]) or "").strip()
        if not project_engineer and pe_map:
            project_engineer = pe_map.get(estimator, "")
        division = normalize_division(row_value(row, ["DIVISION", "Division"])) or division_from_estimator(estimator)
        # Read the FOLLOW UPS sheet's STATUS so "PENDING AWARD" rows can be
        # treated as awarded for goal/Race math (per business decision: pending
        # awards are functionally booked).
        status_raw = str(row_value(row, ["STATUS", "Status"]) or "").strip()
        is_pending_award = status_raw.upper() == "PENDING AWARD"
        outcome = "Awarded" if is_pending_award else "Pending"
        date_awarded = bid_date if is_pending_award else None
        awarded_amount = bid_amount if is_pending_award else 0
        results.append({
            "estNumber": str(est_number).strip(),
            "projectName": str(project).strip(),
            "clientGc": str(row_value(row, ["CLIENT", "Client"]) or "").strip(),
            "contact": "",
            "location": "",
            "scope": "",
            "jobWalk": str(row_value(row, ["JOB WALK", "Job Walk"]) or "").strip(),
            "estimator": estimator,
            "projectEngineer": project_engineer,
            "division": division,
            "bidDueDate": bid_date,
            "bidAmount": bid_amount,
            "awardedAmount": awarded_amount,
            "itemDueMod": safe_number(row_value(row, ["ITEM DUE MOD", "ITEM DUE MOD."])),
            "bidDueTime": str(row_value(row, ["TIME DUE", "BID TIME", "DUE TIME", "Time Due"]) or "").strip(),
            "setup": coerce_bool(row_value(row, ["SETUP"])),
            "documentsReview": coerce_bool(row_value(row, ["DOCUMENTS REVIEW"])),
            "requestQuotes": coerce_bool(row_value(row, ["REQUEST QUOTES"])),
            "takeoffFunding": coerce_bool(row_value(row, ["TAKEOFF & FUNDING", "TAKEOFF AND FUNDING"])),
            "addendaReceived": safe_number(row_value(row, ["ADDENDA RECEIVED"])),
            "addendaConfirmed": safe_number(row_value(row, ["ADDENDA CONFIRMED"])),
            "finalizeNumbers": coerce_bool(row_value(row, ["FINALIZE NUMBERS"])),
            "finalBidCheck": coerce_bool(row_value(row, ["FINAL BID CHECK"])),
            "status": "Sent",
            "outcome": outcome,
            "notes": "",
            "dateSent": bid_date,
            "followUpDate": add_days_iso(bid_date, 60),
            "dateAwarded": date_awarded,
            "awardPending": is_pending_award,
            "followUpStatus": status_raw,
        })
    return results


def parse_archive(df, allowed, pe_map=None):
    results = []
    for _, row in df.iterrows():
        estimator = str(row_value(row, ["ESTIMATOR", "Estimator"]) or "").strip()
        if not estimator or not estimator_allowed(estimator, allowed):
            continue
        project = row_value(row, ["PROJECT/ESTIMATE NAME", "PROJECT/ESTIMATE  NAME", "Project"])
        est_number = row_value(row, ["EST #", "EST#", "Est #"])
        bid_date = excel_date_to_iso(row_value(row, ["BID DATE", "Bid Date"]))
        if not project or not est_number or not bid_date:
            continue
        status_raw = row_value(row, ["STATUS", "Status"]) or ""
        outcome = outcome_from_status(status_raw)
        bid_amount = safe_number(row_value(row, ["PRICE", "Bid Amount", "BID AMOUNT"]))
        awarded_amount = bid_amount if outcome == "Awarded" else 0
        project_engineer = str(row_value(row, ["PE", "Project Engineer", "Project Eng"]) or "").strip()
        if not project_engineer and pe_map:
            project_engineer = pe_map.get(estimator, "")
        division = normalize_division(row_value(row, ["DIVISION", "Division"])) or division_from_estimator(estimator)
        results.append({
            "estNumber": str(est_number).strip(),
            "projectName": str(project).strip(),
            "clientGc": str(row_value(row, ["CLIENT", "Client"]) or "").strip(),
            "contact": "",
            "location": "",
            "scope": "",
            "jobWalk": str(row_value(row, ["JOB WALK", "Job Walk"]) or "").strip(),
            "estimator": estimator,
            "projectEngineer": project_engineer,
            "division": division,
            "bidDueDate": bid_date,
            "bidAmount": bid_amount,
            "awardedAmount": awarded_amount,
            "itemDueMod": safe_number(row_value(row, ["ITEM DUE MOD", "ITEM DUE MOD."])),
            "bidDueTime": str(row_value(row, ["TIME DUE", "BID TIME", "DUE TIME", "Time Due"]) or "").strip(),
            "setup": coerce_bool(row_value(row, ["SETUP"])),
            "documentsReview": coerce_bool(row_value(row, ["DOCUMENTS REVIEW"])),
            "requestQuotes": coerce_bool(row_value(row, ["REQUEST QUOTES"])),
            "takeoffFunding": coerce_bool(row_value(row, ["TAKEOFF & FUNDING", "TAKEOFF AND FUNDING"])),
            "addendaReceived": safe_number(row_value(row, ["ADDENDA RECEIVED"])),
            "addendaConfirmed": safe_number(row_value(row, ["ADDENDA CONFIRMED"])),
            "finalizeNumbers": coerce_bool(row_value(row, ["FINALIZE NUMBERS"])),
            "finalBidCheck": coerce_bool(row_value(row, ["FINAL BID CHECK"])),
            "status": "Archived",
            "outcome": outcome,
            "notes": "",
            "dateSent": bid_date,
            "followUpDate": add_days_iso(bid_date, 60),
            "dateAwarded": bid_date if outcome == "Awarded" else None,
        })
    return results


def extract_estimators(*dfs):
    names = set()
    for df in dfs:
        if df is None or df.empty:
            continue
        col = None
        if "ESTIMATOR" in df.columns:
            col = "ESTIMATOR"
        elif "Estimator" in df.columns:
            col = "Estimator"
        if not col:
            continue
        for val in df[col].dropna().astype(str):
            name = val.strip()
            if name:
                names.add(name)
    return sorted(names, key=lambda s: s.lower())


def build_pe_map(*dfs):
    counts = {}
    for df in dfs:
        if df is None or df.empty:
            continue
        est_col = "ESTIMATOR" if "ESTIMATOR" in df.columns else ("Estimator" if "Estimator" in df.columns else None)
        pe_col = "PE" if "PE" in df.columns else ("Project Engineer" if "Project Engineer" in df.columns else None)
        if not est_col or not pe_col:
            continue
        for _, row in df.iterrows():
            est = str(row.get(est_col, "") or "").strip()
            pe = str(row.get(pe_col, "") or "").strip()
            if not est or not pe:
                continue
            counts.setdefault(est, {})
            counts[est][pe] = counts[est].get(pe, 0) + 1
    mapping = {}
    for est, pe_counts in counts.items():
        best = None
        best_count = -1
        for pe, c in pe_counts.items():
            if c > best_count:
                best = pe
                best_count = c
        if best:
            mapping[est] = best
    return mapping


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--no-filter", action="store_true")
    parser.add_argument("--html", default="index.html")
    parser.add_argument("--skip-core-update", action="store_true")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    bids_df = pd.read_excel(excel_path, sheet_name="BIDS", header=7, engine="openpyxl")
    follow_df = pd.read_excel(excel_path, sheet_name="FOLLOW UPS", header=3, engine="openpyxl")
    archive_df = pd.read_excel(excel_path, sheet_name="ARCHIVE", header=1, engine="openpyxl")

    # Pull Excel-cell hyperlinks from column B (PROJECT/ESTIMATE NAME).
    # pandas drops these — only openpyxl preserves the cell.hyperlink attr.
    # Built once and merged into each record by est_number below.
    documents_url_by_est = _build_documents_url_map(excel_path)
    print(f"Found {len(documents_url_by_est)} bid name hyperlinks (OneDrive / Dropbox / plan-room).")

    excel_estimators = extract_estimators(bids_df, follow_df, archive_df)
    if not args.skip_core_update and excel_estimators:
        update_core_estimators(args.html, excel_estimators)

    allowed = None if args.no_filter else (excel_estimators or load_core_estimators(args.html))
    pe_map = build_pe_map(bids_df, follow_df, archive_df)
    division_map = build_estimator_division_map([(archive_df, 2), (follow_df, 1)])

    bids = []
    def apply_division_map(rows):
        for row in rows:
            if not row.get("division") and row.get("estimator"):
                est = str(row.get("estimator", "")).strip().upper()
                row["division"] = division_map.get(est, "") or division_from_estimator(est)
            # Attach the OneDrive/Dropbox/plan-room link if the BID LIST
            # cell had an Excel hyperlink. Surfaces on the BAY Bid List
            # page as a clickable project name.
            est_num = str(row.get("estNumber", "")).strip().upper()
            url = documents_url_by_est.get(est_num)
            if url:
                row["documentsUrl"] = url
        return rows

    bids.extend(apply_division_map(parse_bids(bids_df, allowed, pe_map)))
    bids.extend(apply_division_map(parse_follow(follow_df, allowed, pe_map)))
    bids.extend(apply_division_map(parse_archive(archive_df, allowed, pe_map)))

    # Deduplicate by estNumber + projectName + status
    seen = set()
    deduped = []
    for b in bids:
        key = (str(b.get("estNumber", "")).upper(), b.get("projectName", ""), b.get("status", ""))
        if key in seen:
            continue
        seen.add(key)
        b["id"] = f"{key[0]}::{b.get('status','')}::{len(deduped)}"
        b["createdAt"] = datetime.utcnow().isoformat() + "Z"
        b["updatedAt"] = datetime.utcnow().isoformat() + "Z"
        deduped.append(b)

    output = {
        "generatedAt": datetime.utcnow().isoformat() + "Z",
        "source": str(excel_path),
        "estimatorDivisions": division_map,
        "bids": deduped,
    }

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(output, indent=2), encoding="utf-8")
    js_path = out_path.with_suffix(".js")
    js_payload = "window.__BIDS_SYNC__ = " + json.dumps(output) + ";\n"
    js_path.write_text(js_payload, encoding="utf-8")
    print(f"Wrote {len(deduped)} bids to {out_path}")


if __name__ == "__main__":
    main()
