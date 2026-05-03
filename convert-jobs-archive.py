"""Convert BUDGET LIST.xlsm > JOBS sheet into the historical job archive.

The JOBS sheet is the master roster of every job the company has ever taken
on. It carries job_id, status, customer, PM/PE/SUP/FM/OPM, dates, contract,
classification, work description, etc. The active BUDGET LIST sheet only
holds in-flight jobs; JOBS is the long tail (archive view) for trend / growth
analysis across years.

We also pull realized cost totals from the COSTS - BUDGS sheet so each
archived job carries its actual spend (labor / material / sub / overhead) as
a roll-up. That gives the archive view a profit estimate per job without
opening the per-job tabs.

Output: window.__STATIC_JOBS_ARCHIVE__ + getStaticJobsArchive() helpers.
"""
import argparse
import json
import re
import warnings
from datetime import datetime
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

JOBS_SHEET = "JOBS"
JOBS_HEADER_ROW = 1
JOBS_DATA_START_ROW = 2

# COSTS - BUDGS layout — three side-by-side blocks: budget rollup, material
# spend, labor spend. We aggregate to one number per (project_id, line).
COSTS_SHEET = "COSTS - BUDGS"
COSTS_HEADER_ROW = 2
COSTS_DATA_START_ROW = 3

JOBS_COL_MAP = {
    "A": "status",                 # A / X / etc.
    "C": "jobId",                  # 2604-SAC, 2599-201
    "D": "description",
    "E": "projectManager",
    "F": "projectEngineer",
    "G": "superintendent",
    "H": "fieldManager",
    "I": "opm",
    "J": "customer",
    "K": "geoArea",                # SAC / BAY (division)
    "L": "cprDirNumber",
    "M": "projectClass",           # EDUCATION / Side Job T&M / Miscellaneous
    "N": "local",
    "O": "awardDate",
    "P": "startDate",
    "Q": "estEndDate",
    "R": "customerPmName",
    "T": "customerFieldContact",
    "V": "estimateNumber",
    "W": "originalContract",
    "X": "bidOrTm",
    "Y": "rate",
    "Z": "grantFunding",
    "AA": "grantFundingAmount",
    "AB": "budgetEntered",
    "AC": "workDescription",
}

# Cost line aliases in COSTS - BUDGS (column B / cost_code_group_id)
COST_LINES = {
    "LABOR": "labor",
    "MATERIAL": "material",
    "SUB": "sub",
    "OVERHEAD": "overhead",
    "OVERSIGHT": "oversight",
}

# CC - ACT is the master cost ledger — every cost transaction across every
# job. Used to compute per-job profit on archived jobs that aren't on the
# active BUDGET LIST anymore.
CC_ACT_SHEET = "CC - ACT"
# Header is on row 2 (row 1 is a super-header strip). Data starts on row 3.
CC_ACT_DATA_START_ROW = 3
# Column A = Project # ; column D = $ amount ; column E = CC Group label.
# (The "Labor Cost" header on column D is misleading — it's the dollar value
# for ANY cost line, not just labor. The CC Group column is what actually
# segregates LABOR / MATERIAL / SUB / OVERHEAD.)
CC_ACT_PROJECT_COL = "A"
CC_ACT_AMOUNT_COL = "D"
CC_ACT_GROUP_COL = "E"

# Map raw CC Group strings into the 4 main buckets we report on. GHST (ghost
# accounts) and OVH/MAT shorthand all fold into their parent line. Anything
# we don't recognize is dropped — we don't want random one-offs polluting
# the totals.
CC_GROUP_BUCKETS = {
    "LABOR": "labor",
    "MATERIAL": "material",
    "MAT": "material",
    "SUB": "sub",
    "OVERHEAD": "overhead",
    "OVH": "overhead",
    "GHST": "overhead",  # ghost / overhead-style allocations
}

NUMERIC_PLACEHOLDERS = {"n/a", "tbd", "no", "yes", ""}


def col_letter_to_index(letter):
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def read_cell(row_tuple, col_letter):
    idx = col_letter_to_index(col_letter)
    if idx >= len(row_tuple):
        return None
    return row_tuple[idx]


def to_number(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s.lower() in NUMERIC_PLACEHOLDERS:
        return None
    cleaned = re.sub(r"[^0-9.\-]", "", s)
    if not cleaned or cleaned in (".", "-", "-."):
        return None
    try:
        return float(cleaned)
    except Exception:
        return None


def num(v):
    n = to_number(v)
    return n if n is not None else 0.0


def clean_str(value):
    if value is None:
        return ""
    return str(value).strip()


def to_iso_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    s = clean_str(value)
    if not s or s.lower() in NUMERIC_PLACEHOLDERS:
        return None
    try:
        return datetime.fromisoformat(s.split(" ")[0]).date().isoformat()
    except Exception:
        return None


def normalize_division(value):
    s = clean_str(value).upper()
    if s.startswith("BAY"):
        return "BAY"
    if s.startswith("SAC"):
        return "SAC"
    return s or "OTHER"


def parse_year(job_id):
    """Job IDs follow YYNN-DIV pattern (e.g. 2604-SAC -> 2026)."""
    s = clean_str(job_id)
    m = re.match(r"^(\d{2})\d{2}", s)
    if not m:
        return None
    yy = int(m.group(1))
    # 19YY for 1900s, 20YY for 2000+. Anything 90-99 = 1990s; 00-89 = 2000s+.
    return 2000 + yy if yy < 90 else 1900 + yy


def load_costs_rollup(wb):
    """Sum LABOR/MATERIAL/SUB/OVERHEAD spend per project_id from COSTS - BUDGS.

    The sheet has three side-by-side blocks; we read the first block (cols
    A-C: project, cost_code_group, orig_est_cost) and the labor/material
    blocks (cols D-F, G-I) which carry actual-cost rollups.
    """
    out = {}
    if COSTS_SHEET not in wb.sheetnames:
        return out
    ws = wb[COSTS_SHEET]
    for row in ws.iter_rows(min_row=COSTS_DATA_START_ROW, values_only=True):
        # Block 1: budget rollup (cols A-C: project, cost_code_group, orig_est_cost)
        proj_b = clean_str(row[0]) if len(row) > 0 else ""
        line_b = clean_str(row[1]).upper() if len(row) > 1 else ""
        budget = to_number(row[2]) if len(row) > 2 else None
        if proj_b and line_b in COST_LINES and budget is not None:
            entry = out.setdefault(proj_b, {"budget": {}, "material": 0.0, "labor": 0.0})
            entry["budget"][COST_LINES[line_b]] = entry["budget"].get(COST_LINES[line_b], 0.0) + budget
        # Block 2: material spend (cols E-F: project, material_cost)
        proj_m = clean_str(row[4]) if len(row) > 4 else ""
        material_cost = to_number(row[5]) if len(row) > 5 else None
        if proj_m and material_cost is not None:
            entry = out.setdefault(proj_m, {"budget": {}, "material": 0.0, "labor": 0.0})
            entry["material"] += material_cost
        # Block 3: labor spend (cols H-I: project, labor_cost)
        proj_l = clean_str(row[7]) if len(row) > 7 else ""
        labor_cost = to_number(row[8]) if len(row) > 8 else None
        if proj_l and labor_cost is not None:
            entry = out.setdefault(proj_l, {"budget": {}, "material": 0.0, "labor": 0.0})
            entry["labor"] += labor_cost
    return out


def load_cc_act_actuals(wb):
    """Roll up CC - ACT (master cost ledger) into actuals per job + line.

    Returns a dict: {project_id: {labor: $, material: $, sub: $, overhead: $}}.
    Also keys the same totals on the bare numeric prefix (e.g. "2151") so we
    can match historic rows whose JOBS.job_id suffix doesn't line up.
    """
    out = {}
    if CC_ACT_SHEET not in wb.sheetnames:
        return out
    ws = wb[CC_ACT_SHEET]
    p_idx = col_letter_to_index(CC_ACT_PROJECT_COL)
    a_idx = col_letter_to_index(CC_ACT_AMOUNT_COL)
    g_idx = col_letter_to_index(CC_ACT_GROUP_COL)
    for row in ws.iter_rows(min_row=CC_ACT_DATA_START_ROW, values_only=True):
        if len(row) <= max(p_idx, a_idx, g_idx):
            continue
        project = clean_str(row[p_idx])
        if not project:
            continue
        group_raw = clean_str(row[g_idx]).upper()
        bucket = CC_GROUP_BUCKETS.get(group_raw)
        if not bucket:
            continue
        amount = to_number(row[a_idx])
        if amount is None:
            continue
        entry = out.setdefault(project, {"labor": 0.0, "material": 0.0, "sub": 0.0, "overhead": 0.0})
        entry[bucket] = entry.get(bucket, 0.0) + float(amount)

        # Mirror to the bare numeric prefix so we can match either form.
        prefix = project.split("-")[0]
        if prefix and prefix != project:
            mirror = out.setdefault(prefix, {"labor": 0.0, "material": 0.0, "sub": 0.0, "overhead": 0.0})
            mirror[bucket] = mirror.get(bucket, 0.0) + float(amount)
    return out


def parse_workbook(wb, *, source_label="(unknown)"):
    """Parse a loaded BUDGET LIST.xlsm workbook -> jobs archive payload.

    Extracted so both the local pipeline (path-based) and the cloud-side
    pipeline (SharePoint share-link bytes) share one parse implementation.
    """
    if JOBS_SHEET not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {JOBS_SHEET}")
    ws = wb[JOBS_SHEET]
    costs = load_costs_rollup(wb)
    cc_actuals = load_cc_act_actuals(wb)

    jobs = []
    for row_tuple in ws.iter_rows(min_row=JOBS_DATA_START_ROW, values_only=True):
        # Skip rows without a job_id.
        job_id = clean_str(read_cell(row_tuple, "C"))
        if not job_id:
            continue

        rec = {}
        for col, key in JOBS_COL_MAP.items():
            raw = read_cell(row_tuple, col)
            if key in {"awardDate", "startDate", "estEndDate"}:
                rec[key] = to_iso_date(raw)
            elif key in {"originalContract", "grantFundingAmount", "rate"}:
                rec[key] = to_number(raw)
            elif key == "geoArea":
                rec[key] = normalize_division(raw)
            else:
                rec[key] = clean_str(raw)

        # Year derived from job_id prefix.
        rec["year"] = parse_year(job_id)

        # Fold in cost rollups when we have them. Match on the full job_id
        # (e.g. "2604-SAC") and on the bare numeric prefix as a fallback for
        # historic jobs whose costs sheet uses a slightly different suffix.
        cost_entry = costs.get(job_id) or costs.get(job_id.split("-")[0])
        if cost_entry:
            b = cost_entry.get("budget", {})
            rec["budgetLabor"] = b.get("labor")
            rec["budgetMaterial"] = b.get("material")
            rec["budgetSub"] = b.get("sub")
            rec["budgetOverhead"] = b.get("overhead")
            total_budget = sum(v for v in (b.get("labor"), b.get("material"), b.get("sub"), b.get("overhead")) if v)
            rec["budgetTotal"] = total_budget if total_budget else None
        else:
            rec["budgetLabor"] = None
            rec["budgetMaterial"] = None
            rec["budgetSub"] = None
            rec["budgetOverhead"] = None
            rec["budgetTotal"] = None

        # Pull master-ledger actuals from CC - ACT (covers ALL cost lines
        # including sub + overhead which COSTS - BUDGS doesn't surface).
        # Falls back to COSTS - BUDGS labor/material when CC - ACT is empty
        # for that job — important for very old archived jobs.
        actuals = cc_actuals.get(job_id) or cc_actuals.get(job_id.split("-")[0])
        if actuals:
            rec["actualLabor"] = actuals.get("labor") or None
            rec["actualMaterial"] = actuals.get("material") or None
            rec["actualSub"] = actuals.get("sub") or None
            rec["actualOverhead"] = actuals.get("overhead") or None
        else:
            rec["actualLabor"] = (cost_entry or {}).get("labor") or None
            rec["actualMaterial"] = (cost_entry or {}).get("material") or None
            rec["actualSub"] = None
            rec["actualOverhead"] = None

        # Total actual cost (zero-fills any missing line) and the profit
        # estimate. We DON'T compute profit when contract is zero/missing
        # or when actuals are too thin to be meaningful (less than 5% of
        # contract suggests we don't really have cost history yet).
        total_actual = sum(num(rec.get(k)) for k in ("actualLabor", "actualMaterial", "actualSub", "actualOverhead"))
        rec["totalActualCost"] = total_actual if total_actual > 0 else None

        contract = num(rec.get("originalContract"))
        if contract > 0 and total_actual > 0 and total_actual >= contract * 0.05:
            rec["estProfit"] = contract - total_actual
            rec["estProfitPct"] = (contract - total_actual) / contract
            rec["profitDataQuality"] = "ok"
        elif contract > 0 and total_actual > 0:
            # Thin actuals — surface but flag.
            rec["estProfit"] = contract - total_actual
            rec["estProfitPct"] = (contract - total_actual) / contract
            rec["profitDataQuality"] = "thin"
        else:
            rec["estProfit"] = None
            rec["estProfitPct"] = None
            rec["profitDataQuality"] = "missing"

        # Convenience: classify status text. "A" looks like Active; anything
        # else flagged as historical (closed/done). Surface raw status too.
        status = rec.get("status", "").upper()
        rec["isActive"] = status == "A"
        rec["statusLabel"] = "Active" if status == "A" else (status or "Closed")

        jobs.append(rec)

    # Aggregates for the dashboard view: rolled up by year + division.
    by_year = {}
    for j in jobs:
        y = j.get("year") or 0
        d = j.get("geoArea") or "OTHER"
        key = (y, d)
        agg = by_year.setdefault(key, {"year": y, "division": d, "jobs": 0, "contractTotal": 0.0,
                                       "actualCost": 0.0, "estProfit": 0.0, "profitJobs": 0})
        agg["jobs"] += 1
        agg["contractTotal"] += num(j.get("originalContract"))
        # Profit aggregates only count jobs with usable cost history. Side
        # jobs and shells without actuals would otherwise inflate profit.
        if j.get("estProfit") is not None and j.get("profitDataQuality") in ("ok", "thin"):
            agg["actualCost"] += num(j.get("totalActualCost"))
            agg["estProfit"] += num(j.get("estProfit"))
            agg["profitJobs"] += 1

    total_profit = sum(num(j.get("estProfit")) for j in jobs if j.get("estProfit") is not None)
    profit_jobs = sum(1 for j in jobs if j.get("estProfit") is not None)

    aggregates = {
        "jobCount": len(jobs),
        "contractTotal": sum(num(j.get("originalContract")) for j in jobs),
        "activeCount": sum(1 for j in jobs if j["isActive"]),
        "estProfitTotal": total_profit,
        "profitJobs": profit_jobs,
        "byYearDivision": sorted(by_year.values(), key=lambda r: (r["year"], r["division"])),
    }

    return {
        "generatedAt": datetime.utcnow().isoformat() + "Z",
        "source": source_label,
        "count": len(jobs),
        "aggregates": aggregates,
        "jobs": jobs,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.excel, data_only=True)
    payload = parse_workbook(wb, source_label=str(args.excel))
    jobs = payload.get("jobs", [])
    payload_json = json.dumps(payload, indent=2)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    js = (
        "// AUTO-GENERATED -- do not edit. Regenerate with fusion-portal/Convert-JobsArchive.ps1\n"
        f"window.__STATIC_JOBS_ARCHIVE__ = {payload_json};\n"
        "window.getStaticJobsArchive = function(){\n"
        "  var p = window.__STATIC_JOBS_ARCHIVE__;\n"
        "  return (p && Array.isArray(p.jobs)) ? p.jobs : [];\n"
        "};\n"
        "window.getStaticJobsArchiveAggregates = function(){\n"
        "  var p = window.__STATIC_JOBS_ARCHIVE__;\n"
        "  return p ? p.aggregates : null;\n"
        "};\n"
    )
    out_path.write_text(js, encoding="utf-8")
    out_path.with_suffix(".json").write_text(payload_json, encoding="utf-8")
    print(f"Wrote {len(jobs)} archived jobs to {out_path}")


if __name__ == "__main__":
    main()
