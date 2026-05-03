"""Convert BUDGET LIST.xlsm > EMPLOYEES into a JS file the Field Employees
page can load. The EMPLOYEES sheet currently has only id + name + UPPER name.
Credential fields (OSHA 10/30, DIR Journeyman card, home local, ITS, forklift,
specialty, etc.) are placeholders until a credentials data source is wired up.

Output: window.getStaticEmployees = function(){ return [...]; };
"""
import argparse
import json
import re
import warnings
from datetime import datetime
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

SHEET_NAME = "EMPLOYEES"
HEADER_ROW = 1
DATA_START_ROW = 2

# Optional secondary sheet with credentials (planned). Will be looked up by
# employee_id when present.
CREDENTIALS_SHEET = "EMPLOYEE_CREDENTIALS"

CRED_FIELDS = [
    "homeLocal",          # 332 / 340 / 6 / 617 / etc.
    "division",           # BAY / SAC
    "title",              # Foreman / Journeyman / Apprentice / Super / OPM
    "osha10",             # date earned or boolean
    "osha30",
    "dirJourneyman",
    "itsCertified",
    "forkliftCertified",
    "scissorLift",
    "boomLift",
    "firstAidCpr",
    "specialty",          # Lighting / Distribution / Communications / etc.
    "phone",
    "email",
    "active",             # true/false
    "notes",
]


def col_letter_to_index(letter):
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch.upper()) - ord("A") + 1)
    return n - 1


def clean_str(value):
    if value is None: return ""
    return str(value).strip()


def normalize_division(value):
    s = clean_str(value).upper()
    if s.startswith("BAY"): return "BAY"
    if s.startswith("SAC"): return "SAC"
    return s


def load_credentials(wb):
    """Optional credentials sheet keyed by employee_id."""
    out = {}
    if CREDENTIALS_SHEET not in wb.sheetnames:
        return out
    ws = wb[CREDENTIALS_SHEET]
    headers = [clean_str(c.value).lower() for c in ws[1]]
    if "employee_id" not in headers:
        return out
    id_idx = headers.index("employee_id")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[id_idx]: continue
        emp_id = str(row[id_idx]).strip()
        rec = {}
        for i, h in enumerate(headers):
            if not h or h == "employee_id": continue
            if i >= len(row): continue
            v = row[i]
            if v is None: continue
            rec[h] = clean_str(v)
        out[emp_id] = rec
    return out


def parse_workbook(wb, *, source_label="(unknown)"):
    """Parse a loaded workbook (PROJECT LIST.xlsm) -> employees payload.

    The EMPLOYEES sheet lives in PROJECT LIST.xlsm (the docstring's
    older "BUDGET LIST.xlsm" reference is outdated -- script just reads
    whatever workbook gets passed in, by sheet name). Extracted so both
    the local pipeline and the cloud-side pipeline share one parse.
    """
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")
    ws = wb[SHEET_NAME]
    creds = load_credentials(wb)

    employees = []
    seen_ids = set()
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row or len(row) < 2: continue
        emp_id_raw = row[0]
        name = clean_str(row[1] if len(row) > 1 else None)
        if not emp_id_raw and not name: continue
        emp_id = str(emp_id_raw).strip() if emp_id_raw is not None else ""
        if not name: continue
        if emp_id in seen_ids: continue
        seen_ids.add(emp_id)
        rec = {
            "id": emp_id or f"name-{re.sub(r'[^a-zA-Z0-9]+','-',name).strip('-').lower()}",
            "employeeId": emp_id,
            "name": name,
            "nameUpper": clean_str(row[2] if len(row) > 2 else None) or name.upper(),
        }
        # Defaults for credential fields
        for f in CRED_FIELDS:
            rec[f] = ""
        rec["active"] = "true"
        # Overlay actual credential data when present
        cred = creds.get(emp_id) or {}
        for k, v in cred.items():
            rec[k] = v
        if rec.get("division"):
            rec["division"] = normalize_division(rec["division"])
        employees.append(rec)

    return {
        "generatedAt": datetime.utcnow().isoformat() + "Z",
        "source": source_label,
        "count": len(employees),
        "credentialFields": CRED_FIELDS,
        "employees": employees,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.excel, data_only=True)
    payload = parse_workbook(wb, source_label=str(args.excel))
    employees = payload.get("employees", [])
    payload_json = json.dumps(payload, indent=2)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    js = (
        "// AUTO-GENERATED -- regenerate with fusion-field-employees/Convert-Employees.ps1\n"
        f"window.__STATIC_EMPLOYEES__ = {payload_json};\n"
        "window.getStaticEmployees = function(){\n"
        "  var p = window.__STATIC_EMPLOYEES__;\n"
        "  return (p && Array.isArray(p.employees)) ? p.employees : [];\n"
        "};\n"
    )
    out_path.write_text(js, encoding="utf-8")
    out_path.with_suffix(".json").write_text(payload_json, encoding="utf-8")
    print(f"Wrote {len(employees)} employees to {out_path}")


if __name__ == "__main__":
    main()
