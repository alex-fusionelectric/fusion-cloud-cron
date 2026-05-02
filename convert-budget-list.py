"""Convert BUDGET LIST.xlsm > BUDGET LIST sheet into a JS file the portal can
load directly.

The BUDGET LIST sheet (the main one — not PM REPORT, which is a SAC-only
working view) is the canonical source for active project budget tracking. It
covers both BAY and SAC and has the raw labor / material / sub / overhead
budget+cost columns that drive variance and profit math.

Output: `window.getStaticBudgetProjects = function(){ return [...]; };`
"""
import argparse
import json
import re
import warnings
from datetime import datetime
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

SHEET_NAME = "BUDGET LIST"
HEADER_ROW = 3
DATA_START_ROW = 4

# PM REPORT (SAC ops working view) has free-text action items in column Z. We
# pull them when present so the portal surfaces the same notes the team already
# discusses in their daily review.
PM_REPORT_SHEET = "PM REPORT"
PM_REPORT_HEADER_ROW = 7
PM_REPORT_DATA_START_ROW = 8
PM_REPORT_NUMBER_COL = "C"   # e.g. "2606-SAC"
PM_REPORT_ACTION_COL = "Z"

# Column letter -> portal field. Headers from row 3 of BUDGET LIST.
COL_MAP = {
    "B": "fullNumber",            # FULL # FOR FD (e.g. "2611-BAY")
    "C": "projectName",
    "D": "jobNumber",
    "G": "overdueFlag",
    "H": "projectManager",
    "I": "projectEngineer",
    "J": "generalForeman",
    "K": "projectForeman",
    "L": "customer",
    "M": "division",              # S/B (BAY / SAC)
    "O": "foundationStatus",      # A = approved, blank/TBD = missing
    "P": "awardDate",
    "Q": "startDate",
    "S": "endDate",
    "W": "originalContract",
    "X": "approvedCos",
    "Y": "pendingCos",
    "AA": "subAndPoTotal",
    "AB": "contractBilled",
    "AC": "contractPayReceived",
    "AE": "jobListStatus",
    "AR": "pctComplete",          # EST COMP % (decimal)
    "AS": "projectedPace",
    "AT": "laborBudget",
    "AU": "laborCost",
    "AV": "subBudget",
    "AW": "subCost",
    "AX": "materialBudget",
    "AY": "materialCost",
    "AZ": "overheadBudget",
    "BA": "overheadCost",
    "BB": "budgetHoursProjected",
    "BC": "completeFlag",
    "BD": "billingPct",
}

NUMERIC_PLACEHOLDERS = {
    "no labor yet", "no labor % entered", "no labor",
    "missing data", "n/a", "tbd", "no", "",
}


def col_letter_to_index(letter):
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch.upper()) - ord("A") + 1)
    return n - 1


def read_cell(row_tuple, col_letter):
    idx = col_letter_to_index(col_letter)
    if idx >= len(row_tuple):
        return None
    return row_tuple[idx]


def to_number(value):
    if value is None: return None
    if isinstance(value, bool): return None
    if isinstance(value, (int, float)): return float(value)
    s = str(value).strip()
    if s.lower() in NUMERIC_PLACEHOLDERS: return None
    cleaned = re.sub(r"[^0-9.\-]", "", s)
    if not cleaned or cleaned in (".", "-", "-."): return None
    try: return float(cleaned)
    except Exception: return None


def num(v):
    """Same as to_number but returns 0 instead of None for arithmetic."""
    n = to_number(v)
    return n if n is not None else 0.0


def clean_str(value):
    if value is None: return ""
    return str(value).strip()


def to_iso_date(value):
    if value is None: return None
    if isinstance(value, datetime): return value.date().isoformat()
    s = clean_str(value)
    if not s or s.lower() in NUMERIC_PLACEHOLDERS: return None
    try: return datetime.fromisoformat(s.split(" ")[0]).date().isoformat()
    except Exception: return None


def normalize_division(value):
    s = clean_str(value).upper()
    if s.startswith("BAY"): return "BAY"
    if s.startswith("SAC"): return "SAC"
    if s.startswith("BOTH"): return "BOTH"
    return s


def project_health(profit_pct):
    if profit_pct is None: return "unknown"
    if profit_pct >= 0.15: return "good"
    if profit_pct >= 0.05: return "watch"
    return "risk"


def project_cost_line(cost, budget, pct_complete, line_kind="generic"):
    """Project the FINAL cost of a line item.

    Reverse-engineered from PM REPORT (the SAC ops-of-record working view) so
    portal totals land close to its $6.46M projected profit.

    LABOR: extrapolate `cost / pct_complete` once past 10% complete. Labor is
    time-based and scales linearly with progress; below 10% the data is too
    noisy and we fall back to budget.

    MATERIAL / OVERHEAD:
      - Past ~50% complete: use actual cost. Materials are typically bought
        early so by midpoint actual ≈ final. Overhead is mostly recognized
        as work proceeds — past midpoint actual is a better estimate than
        budget. If actual is zero, fall back to budget.
      - Below 50% complete: use max(actual, budget). Conservative — picks up
        runaways once actuals exceed budget but doesn't extrapolate to
        unrealistic numbers when we're early and just front-loaded a buy.

    SUB: contractually committed; always use max(actual, budget).
    """
    cost = max(0.0, float(cost or 0))
    budget = max(0.0, float(budget or 0))
    if line_kind == "labor":
        if pct_complete is not None and pct_complete > 0.1 and cost > 0:
            return cost / pct_complete
        return budget if budget > 0 else cost
    if line_kind == "sub":
        return max(cost, budget)
    # material / overhead
    if pct_complete is not None and pct_complete >= 0.5:
        return cost if cost > 0 else budget
    return max(cost, budget)


def load_pm_report_actions(wb):
    """Build a map: job_number -> action item text from the PM REPORT sheet.
    PM REPORT is a SAC-only working view; for any project listed there we
    surface its 'ACTION ITEM(S)' string verbatim."""
    out = {}
    if PM_REPORT_SHEET not in wb.sheetnames:
        return out
    ws = wb[PM_REPORT_SHEET]
    num_idx = col_letter_to_index(PM_REPORT_NUMBER_COL)
    act_idx = col_letter_to_index(PM_REPORT_ACTION_COL)
    for row_tuple in ws.iter_rows(min_row=PM_REPORT_DATA_START_ROW, values_only=True):
        if len(row_tuple) <= max(num_idx, act_idx): continue
        full = clean_str(row_tuple[num_idx])
        action = clean_str(row_tuple[act_idx])
        if not full or not action or full == "-":
            continue
        # Strip "-BAY" / "-SAC" suffix to get just the job number for matching.
        m = re.match(r"^\s*(\S+?)\s*-", full)
        job = m.group(1) if m else full
        out[job.strip()] = action
    return out


def derive_action_items(rec):
    """Synthesize action notes when PM REPORT doesn't have an explicit one
    (BAY projects, or SAC jobs PM REPORT doesn't track yet). Each note is a
    short phrase the team can scan."""
    items = []
    pct = rec.get("pctComplete")
    labor_budget = num(rec.get("laborBudget"))
    labor_cost = num(rec.get("laborCost"))
    material_budget = num(rec.get("materialBudget"))
    material_cost = num(rec.get("materialCost"))
    overhead_budget = num(rec.get("overheadBudget"))
    overhead_cost = num(rec.get("overheadCost"))
    pending_cos = num(rec.get("pendingCos"))
    foundation = clean_str(rec.get("foundationStatus"))
    proj_labor = num(rec.get("estFinalLaborCost"))
    proj_material = num(rec.get("estFinalMaterialCost"))

    # Over-budget flags (5%+ over)
    if labor_budget > 0 and proj_labor > labor_budget * 1.05:
        items.append("Labor over budget")
    if material_budget > 0 and proj_material > material_budget * 1.05:
        items.append("Material over budget")
    if overhead_budget > 0 and overhead_cost > overhead_budget * 1.05:
        items.append("Overhead over budget")
    # Pending COs sitting around
    if pending_cos > 0:
        items.append(f"${int(pending_cos):,} in pending CO's to approve")
    # Foundation status missing
    f_norm = foundation.upper()
    if not f_norm or f_norm in {"TBD", "N/A", "MISSING"}:
        items.append("Foundation info missing")
    # No labor % entered yet despite labor cost showing
    if labor_cost > 0 and (pct is None or pct == 0):
        items.append("Add labor %'s on tab")
    return "; ".join(items)


def parse_workbook(wb, *, source_label="(unknown)"):
    """Parse a loaded BUDGET LIST.xlsm workbook into the portal payload dict.

    Extracted so both the local pipeline (which loads the workbook from
    a path) and the cloud-side pipeline (which loads it from the bytes of
    a SharePoint share-link download) share one parse implementation. If
    you change the parse logic here it flows through both paths -- which
    is the whole point.
    """
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")
    ws = wb[SHEET_NAME]
    pm_actions = load_pm_report_actions(wb)

    projects = []
    for row_tuple in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        # Skip rows without a project name + full number — those are blank
        # template/spacer rows.
        full_number = clean_str(read_cell(row_tuple, "B"))
        project_name = clean_str(read_cell(row_tuple, "C"))
        if not full_number or not project_name:
            continue

        # Pull every mapped column.
        rec = {}
        for col, key in COL_MAP.items():
            raw = read_cell(row_tuple, col)
            if key in {"fullNumber", "projectName", "jobNumber",
                       "projectManager", "projectEngineer", "generalForeman",
                       "projectForeman", "customer", "jobListStatus",
                       "completeFlag", "projectedPace", "foundationStatus"}:
                rec[key] = clean_str(raw)
            elif key in {"awardDate", "startDate", "endDate"}:
                rec[key] = to_iso_date(raw)
            elif key == "division":
                rec[key] = normalize_division(raw)
            else:
                rec[key] = to_number(raw)

        # NOTE: previously dropped rows where completeFlag (BC) was YES/DONE.
        # That removed jobs the user still tracks in the active BUDGET LIST
        # sheet (e.g. 2511 IRVINGTON PARK RESTROOM, 2453 300P PET CT — both at
        # 100% complete but flagged for variance review). The Excel sheet's
        # presence/absence of a row is now the source of truth; if a job needs
        # to disappear from the portal, remove it from the sheet.

        # Project must have at least one positive budget line to be real;
        # otherwise it's a placeholder / not-yet-budgeted shell.
        has_any_budget = any(
            (rec.get(k) or 0) > 0
            for k in ("laborBudget", "materialBudget", "subBudget", "overheadBudget", "originalContract")
        )
        if not has_any_budget:
            continue

        # Derived fields.
        labor_budget = num(rec.get("laborBudget"))
        labor_cost = num(rec.get("laborCost"))
        material_budget = num(rec.get("materialBudget"))
        material_cost = num(rec.get("materialCost"))
        sub_budget = num(rec.get("subBudget"))
        sub_cost = num(rec.get("subCost"))
        overhead_budget = num(rec.get("overheadBudget"))
        overhead_cost = num(rec.get("overheadCost"))
        original_contract = num(rec.get("originalContract"))
        approved_cos = num(rec.get("approvedCos"))
        pending_cos = num(rec.get("pendingCos"))

        contract_total = original_contract + approved_cos + pending_cos
        total_budget = labor_budget + material_budget + sub_budget + overhead_budget
        pct_complete = rec.get("pctComplete")

        # Project each cost line forward to its likely FINAL value. The raw
        # "*_COST" columns in BUDGET LIST are spend-to-date, not projections,
        # so just summing them understates total cost on any in-flight job.
        # PM REPORT does this projection for SAC; we replicate it for all jobs
        # (BAY + SAC) so the numbers track reality across both yards.
        proj_labor = project_cost_line(labor_cost, labor_budget, pct_complete, "labor")
        proj_material = project_cost_line(material_cost, material_budget, pct_complete, "material")
        proj_sub = project_cost_line(sub_cost, sub_budget, pct_complete, "sub")
        proj_overhead = project_cost_line(overhead_cost, overhead_budget, pct_complete, "overhead")
        est_total_cost = proj_labor + proj_material + proj_sub + proj_overhead

        final_budget_diff = total_budget - est_total_cost  # +ve = under budget
        est_profit = contract_total - est_total_cost
        est_profit_pct = (est_profit / contract_total) if contract_total > 0 else None
        labor_diff = labor_budget - proj_labor
        material_diff = material_budget - proj_material
        if pct_complete is None and total_budget > 0:
            pct_complete = est_total_cost / total_budget if total_budget > 0 else None

        rec["contractTotal"] = contract_total
        rec["totalBudget"] = total_budget
        rec["estTotalCost"] = est_total_cost
        rec["finalBudgetDifference"] = final_budget_diff
        rec["budgetDifferencePct"] = (final_budget_diff / total_budget) if total_budget > 0 else None
        rec["estProfit"] = est_profit
        rec["estProfitPct"] = est_profit_pct
        rec["estFinalLaborCost"] = proj_labor or None
        rec["estFinalLaborDifference"] = labor_diff if labor_budget else None
        rec["laborPct"] = (proj_labor / labor_budget) if labor_budget > 0 else None
        rec["estFinalMaterialCost"] = proj_material or None
        rec["estFinalMaterialDifference"] = material_diff if material_budget else None
        rec["materialPct"] = (proj_material / material_budget) if material_budget > 0 else None
        rec["estFinalSubCost"] = proj_sub or None
        rec["estFinalOverheadCost"] = proj_overhead or None
        # Keep the raw spend-to-date around for tooltips / debugging.
        rec["actualLaborCost"] = labor_cost
        rec["actualMaterialCost"] = material_cost
        rec["actualSubCost"] = sub_cost
        rec["actualOverheadCost"] = overhead_cost
        rec["pctComplete"] = pct_complete
        rec["health"] = project_health(est_profit_pct)
        # Pull the PM REPORT note when it exists, otherwise synthesize from
        # data heuristics so BAY jobs (which PM REPORT doesn't cover) still
        # show actionable notes.
        job_no = clean_str(rec.get("jobNumber"))
        pm_note = pm_actions.get(job_no, "")
        derived = derive_action_items(rec)
        if pm_note and derived:
            rec["actionItems"] = f"{pm_note} · {derived}"
        elif pm_note:
            rec["actionItems"] = pm_note
        else:
            rec["actionItems"] = derived
        rec["totalBilled"] = num(rec.get("contractBilled"))
        rec["overUnderBilling"] = None
        rec["pendingCosToApprove"] = pending_cos

        rec["id"] = f"budget-{rec.get('jobNumber','')}".lower().replace(" ", "-")
        projects.append(rec)

    # Aggregates.
    def safe_sum(rows, key):
        return sum(float(r.get(key) or 0) for r in rows)

    aggregates = {
        "projectCount": len(projects),
        "totalContract": safe_sum(projects, "contractTotal"),
        "totalBilled": safe_sum(projects, "totalBilled"),
        "totalEstProfit": safe_sum(projects, "estProfit"),
        "totalBudget": safe_sum(projects, "totalBudget"),
        "totalEstTotalCost": safe_sum(projects, "estTotalCost"),
        "totalLaborBudget": safe_sum(projects, "laborBudget"),
        "totalLaborEstFinal": safe_sum(projects, "estFinalLaborCost"),
        "totalMaterialBudget": safe_sum(projects, "materialBudget"),
        "totalMaterialEstFinal": safe_sum(projects, "estFinalMaterialCost"),
        "byHealth": {
            "good": sum(1 for p in projects if p["health"] == "good"),
            "watch": sum(1 for p in projects if p["health"] == "watch"),
            "risk": sum(1 for p in projects if p["health"] == "risk"),
            "unknown": sum(1 for p in projects if p["health"] == "unknown"),
        },
        "byDivision": {
            "BAY": sum(1 for p in projects if p["division"] == "BAY"),
            "SAC": sum(1 for p in projects if p["division"] == "SAC"),
            "OTHER": sum(1 for p in projects if p["division"] not in {"BAY", "SAC"}),
        },
    }

    return {
        "generatedAt": datetime.utcnow().isoformat() + "Z",
        "source": source_label,
        "count": len(projects),
        "aggregates": aggregates,
        "projects": projects,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.excel, data_only=True)
    payload = parse_workbook(wb, source_label=str(args.excel))
    payload_json = json.dumps(payload, indent=2)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    js = (
        "// AUTO-GENERATED -- do not edit. Regenerate with fusion-portal/Convert-BudgetList.ps1\n"
        f"window.__STATIC_BUDGET__ = {payload_json};\n"
        "window.getStaticBudgetProjects = function(){\n"
        "  var p = window.__STATIC_BUDGET__;\n"
        "  return (p && Array.isArray(p.projects)) ? p.projects : [];\n"
        "};\n"
        "window.getStaticBudgetAggregates = function(){\n"
        "  var p = window.__STATIC_BUDGET__;\n"
        "  return p ? p.aggregates : null;\n"
        "};\n"
    )
    out_path.write_text(js, encoding="utf-8")
    out_path.with_suffix(".json").write_text(payload_json, encoding="utf-8")
    print(f"Wrote {len(payload['projects'])} budget records to {out_path}")


if __name__ == "__main__":
    main()
