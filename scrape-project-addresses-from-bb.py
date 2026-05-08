#!/usr/bin/env python3
"""scrape-project-addresses-from-bb.py

Walks each active project's Dropbox folder, finds the BID BREAKDOWN xlsm,
and pulls the project address out of it heuristically (BB layouts aren't
perfectly consistent across years/divisions). Writes into
public.project_locations_cloud where the geocoder cron picks it up next.

Heuristic — looks for any cell whose value contains one of the address
label tokens (PROJECT ADDRESS, JOB ADDRESS, SITE ADDRESS, LOCATION,
JOBSITE), then takes the cell immediately to the RIGHT in the same row;
falls back to the cell BELOW if right is empty. Final value is validated:
must contain at least one digit AND a comma to be considered an address.

Required env: SUPABASE_SERVICE_KEY, DROPBOX_REFRESH_TOKEN, DROPBOX_APP_KEY,
              DROPBOX_APP_SECRET
"""
from __future__ import annotations
import io
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone

import dropbox  # type: ignore
import openpyxl  # type: ignore
from dropbox import common as dropbox_common  # type: ignore
from dropbox.exceptions import ApiError, AuthError  # type: ignore

SUPABASE_URL = "https://dltuvsdwrujjsmiotaxy.supabase.co"

# Dropbox layout for ACTIVE projects (per real path Alex confirmed
# 2026-05-08):
#   /Fusion Electric Folder/01- PROJECTS/BAY PROJECTS/<full_number>-<NAME>/
#                                                    /01- <short#> CONTRACTS & CO'S/
#                                                                       BID BREAKDOWN*.xlsm
# Example:
#   /Fusion Electric Folder/01- PROJECTS/BAY PROJECTS/2611-BAY-FUSD CAB, GLEN, PATT HVAC/
#       01- 2611 CONTRACTS & CO'S/BID BREAKDOWN - V10.49 - FUSD CAB GLEN PATT ES HVAC.xlsm
PROJECT_ROOTS = {
    "BAY": "/Fusion Electric Folder/01- PROJECTS/BAY PROJECTS",
    "SAC": "/Fusion Electric Folder/01- PROJECTS/SAC PROJECTS",
}
# Filename patterns that identify the BB. We try both common spellings.
BB_FILE_PATTERNS = [
    re.compile(r"^BID\s*BREAKDOWN.*\.xlsm$", re.IGNORECASE),
    re.compile(r"^BB.*\.xlsm$", re.IGNORECASE),
    re.compile(r"^.*BID\s*BREAK.*\.xlsm$", re.IGNORECASE),
]
# Cell-label tokens that flag the address row
ADDRESS_LABELS = (
    "PROJECT ADDRESS", "JOB ADDRESS", "SITE ADDRESS",
    "JOBSITE ADDRESS", "JOB SITE", "JOBSITE", "LOCATION",
    "ADDRESS",  # last so the more-specific labels win
)
# Sanity check that a candidate value looks like an address
ADDR_RE = re.compile(r"^[\d].*[\d]")  # starts with digit, has another digit
# A useful address has at least a number + comma (street + city/state)
LOOKS_LIKE_ADDR = re.compile(r"\d.*,")

# Limit so a one-shot run doesn't try to read every project ever; the cron
# will keep up with new ones. Increase for one-time backfills.
MAX_PROJECTS = int(os.environ.get("MAX_PROJECTS", "200"))


def _service_key() -> str:
    k = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()
    if not k:
        raise SystemExit("SUPABASE_SERVICE_KEY env var required.")
    return k


def _sb(method, path, body=None, extra=None, timeout=30):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {
        "apikey": _service_key(),
        "Authorization": f"Bearer {_service_key()}",
        "content-type": "application/json",
    }
    if extra:
        headers.update(extra)
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, method=method, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.status, r.read()
    except urllib.error.HTTPError as e:
        return e.code, e.read()


def dropbox_client() -> dropbox.Dropbox:
    """Configure for the TEAM root namespace — Fusion folders live there,
    not in the auth user's personal namespace. Mirrors cloud-sync-prequal."""
    rt = os.environ.get("DROPBOX_REFRESH_TOKEN")
    ak = os.environ.get("DROPBOX_APP_KEY")
    asec = os.environ.get("DROPBOX_APP_SECRET")
    if not (rt and ak and asec):
        raise SystemExit("DROPBOX_REFRESH_TOKEN/APP_KEY/APP_SECRET env vars required.")
    dbx = dropbox.Dropbox(oauth2_refresh_token=rt, app_key=ak, app_secret=asec, timeout=60)
    try:
        acct = dbx.users_get_current_account()
    except AuthError as e:
        raise SystemExit(f"Dropbox auth failed: {e}")
    ri = acct.root_info
    root_ns = getattr(ri, "root_namespace_id", None)
    home_ns = getattr(ri, "home_namespace_id", None)
    if root_ns and root_ns != home_ns:
        dbx = dbx.with_path_root(dropbox_common.PathRoot.root(root_ns))
        print(f"Using team root namespace {root_ns}")
    return dbx


def fetch_active_projects() -> list[dict]:
    """Pull active projects from projects_cloud (jobListStatus CURRENT/READY/NO LABOR YET).
    Skips ones already in project_locations_cloud. Returns full_number plus
    the originalEstNumber (e.g. "26-228") so we can locate the project's
    EST folder under /02- ESTIMATING/002- SENT ESTIMATES."""
    st, body = _sb("GET", "projects_cloud?select=full_number,project_name,division,payload&full_number=not.is.null")
    main = []
    if st == 200:
        for r in json.loads(body):
            pl = r.get("payload") or {}
            jls = (pl.get("jobListStatus") or "").upper()
            if jls not in ("CURRENT", "READY TO CLOSE", "NO LABOR YET"):
                continue
            main.append({
                "full_number": r["full_number"],
                "project_name": r.get("project_name") or "",
                "division": (r.get("division") or "").upper(),
                "est_number": (pl.get("originalEstNumber") or pl.get("estNumber") or "").strip(),
            })

    # 2. Skip rows already in project_locations_cloud
    st, body = _sb("GET", "project_locations_cloud?select=full_number")
    have = set()
    if st == 200:
        for r in json.loads(body):
            if r.get("full_number"):
                have.add(r["full_number"])
    return [p for p in main if p["full_number"] not in have][:MAX_PROJECTS]


def list_dbx_folder(dbx: dropbox.Dropbox, path: str) -> list:
    """Returns Dropbox entries for `path`, paging through if needed."""
    entries = []
    try:
        result = dbx.files_list_folder(path, recursive=False)
        entries.extend(result.entries)
        while result.has_more:
            result = dbx.files_list_folder_continue(result.cursor)
            entries.extend(result.entries)
    except ApiError as e:
        # Folder doesn't exist or no access — silently fall through
        if "not_found" not in str(e):
            print(f"  [warn] list {path}: {e}")
    return entries


def find_project_folder(dbx, full_number: str, project_name: str, division: str) -> str | None:
    """Project folders are named "<full_number>-<NAME>" (e.g.
    "2611-BAY-FUSD CAB, GLEN, PATT HVAC") under the division root.
    Match on the full_number prefix."""
    root = PROJECT_ROOTS.get(division.upper())
    if not root:
        return None
    needle = full_number.upper() + "-"   # trailing dash forces exact prefix
    for entry in list_dbx_folder(dbx, root):
        if not isinstance(entry, dropbox.files.FolderMetadata): continue
        if entry.name.upper().startswith(needle):
            return entry.path_lower
    return None


def find_bb_file(dbx, folder_path: str) -> str | None:
    """The BB lives one level deep, in the "01- <short#> CONTRACTS & CO'S"
    subfolder of the project folder. We walk the project folder's
    immediate children, find any subfolder containing "CONTRACTS" + check
    inside for a "BID BREAKDOWN*.xlsm" filename. Falls back to scanning
    the project folder root + every other subfolder if the convention
    isn't followed for a particular project."""
    def files_in(path: str) -> list[str]:
        out = []
        for entry in list_dbx_folder(dbx, path):
            if isinstance(entry, dropbox.files.FileMetadata):
                for pat in BB_FILE_PATTERNS:
                    if pat.match(entry.name):
                        out.append(entry.path_lower)
                        break
        return out

    # 1. Preferred: walk into "*CONTRACTS*" subfolder
    contracts_subs = []
    other_subs = []
    for entry in list_dbx_folder(dbx, folder_path):
        if isinstance(entry, dropbox.files.FolderMetadata):
            if "CONTRACTS" in entry.name.upper():
                contracts_subs.append(entry.path_lower)
            else:
                other_subs.append(entry.path_lower)

    candidates = []
    for sub in contracts_subs:
        candidates.extend(files_in(sub))
    if not candidates:
        # 2. Fallback: project folder root
        candidates.extend(files_in(folder_path))
    if not candidates:
        # 3. Last resort: any other subfolder
        for sub in other_subs:
            candidates.extend(files_in(sub))
            if candidates: break
    # Pick the shortest path (likely the canonical, non-archived copy)
    candidates.sort(key=len)
    return candidates[0] if candidates else None


def extract_address_from_bb(dbx, bb_path: str) -> str | None:
    """Download BB, look for an address in the ESTIMATE SETUP sheet (or
    whatever the first sheet is). We look for any cell containing one of
    the ADDRESS_LABELS, then take the cell to its right; falls back to
    cell below if right is empty."""
    try:
        _, resp = dbx.files_download(bb_path)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        return None
    # Prefer the ESTIMATE SETUP sheet (where we saw "PROJECT ADDRESS" in
    # row 11). Fall back to the first sheet.
    for sheet_name in wb.sheetnames:
        if "ESTIMATE SETUP" in sheet_name.upper() or "SETUP" in sheet_name.upper():
            ws = wb[sheet_name]
            break
    else:
        ws = wb[wb.sheetnames[0]]

    # Walk the first ~40 rows / 12 cols looking for an address label.
    # BBs we've seen put PROJECT ADDRESS in column A and the value in B,
    # but we don't assume it.
    grid = []
    for row in ws.iter_rows(min_row=1, max_row=40, max_col=12, values_only=True):
        grid.append(list(row) if row else [])
    for r_idx, row in enumerate(grid):
        for c_idx, val in enumerate(row):
            if not val: continue
            sval = str(val).strip().upper()
            if not any(lab in sval for lab in ADDRESS_LABELS): continue
            # Same row, next col
            if c_idx + 1 < len(row):
                cand = row[c_idx + 1]
                if cand and str(cand).strip():
                    s = str(cand).strip()
                    if LOOKS_LIKE_ADDR.search(s):
                        return s
            # Below
            if r_idx + 1 < len(grid):
                below = grid[r_idx + 1]
                if c_idx < len(below) and below[c_idx]:
                    s = str(below[c_idx]).strip()
                    if LOOKS_LIKE_ADDR.search(s):
                        return s
    return None


def upsert_location(full_number: str, address: str) -> None:
    iso = datetime.now(timezone.utc).isoformat(timespec="seconds")
    body = [{
        "full_number": full_number,
        "address": address[:500],
        "source": "bb_scrape",
        "updated_at": iso,
    }]
    st, resp = _sb("POST", "project_locations_cloud",
                   body=body, extra={"Prefer": "resolution=merge-duplicates,return=minimal"})
    if st not in (200, 201, 204):
        print(f"  [warn] upsert {full_number} HTTP {st}: {resp[:200]!r}")


def main():
    targets = fetch_active_projects()
    print(f"=== scrape-project-addresses-from-bb ({len(targets)} active without addresses) ===")
    if not targets:
        print("Nothing to do.")
        return
    dbx = dropbox_client()

    found = 0
    for i, p in enumerate(targets, 1):
        fn = p["full_number"]
        div = (p.get("division") or "").upper()
        print(f"[{i}/{len(targets)}] {fn:14s} {div:3s} {p['project_name'][:35]:35s}", end=" ")
        folder = find_project_folder(dbx, fn, p["project_name"], div)
        if not folder:
            print("→ folder not found")
            continue
        bb = find_bb_file(dbx, folder)
        if not bb:
            print("→ BB not found")
            continue
        addr = extract_address_from_bb(dbx, bb)
        if not addr:
            print("→ no address in BB")
            continue
        print(f"→ {addr[:60]}")
        upsert_location(fn, addr)
        found += 1
    print(f"\nDone — {found} address(es) written.")


if __name__ == "__main__":
    main()
