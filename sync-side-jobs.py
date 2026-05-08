#!/usr/bin/env python3
"""sync-side-jobs.py

Parses per-year side-job tabs out of PROJECT LIST.xlsm (2699B, 2699S, 2599B,
2599S, 2499B, 2499S, etc.) and syncs only CURRENT-status jobs into
public.side_jobs_cloud. Inactive states (COMPLETE, PAID, NO BILLING,
INVOICE SENT, blank) are excluded at write time so the field-panel picker
never has to filter — the table itself is always "what's open right now."

Tab layout (per inspection 2026-05-08):
  Header row     = 13
  First data row = 14
  Columns we care about:
    B = Project name
    C = Job # ("2699-104")
    D = Suffix only ("104")
    E = EST #
    F = Award date
    G = Customer
    H = Contact
    I = Project manager
    J = Foreman
    M = JOB STATUS    <-- the filter column
    O = Bonus status
    S = Work description / notes
    Y = Contract OG (amount)

Required env: PROJECT_LIST_URL (or local path), SUPABASE_SERVICE_KEY
"""
from __future__ import annotations
import json
import os
import re
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone

import openpyxl  # type: ignore

SUPABASE_URL = "https://dltuvsdwrujjsmiotaxy.supabase.co"
TABLE = "side_jobs_cloud"

# Tab name → (year, division). Add to this map as new years roll over.
# The B/S suffix is BAY/SAC — Alex's existing convention.
TAB_MAP = {
    "2699B": (2026, "BAY"), "2699S": (2026, "SAC"),
    "2599B": (2025, "BAY"), "2599S": (2025, "SAC"),
    "2499B": (2024, "BAY"), "2499S": (2024, "SAC"),
    "2399B": (2023, "BAY"), "2399S": (2023, "SAC"),
}

# Only these statuses are kept. Anything else (COMPLETE, NO BILLING,
# INVOICE SENT, blank) is excluded so the table stays "open jobs only."
ACTIVE_STATUSES = {"CURRENT"}

HEADER_ROW = 13
DATA_START_ROW = 14


def _service_key() -> str:
    k = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()
    if not k:
        raise SystemExit("SUPABASE_SERVICE_KEY env var required.")
    return k


def _sb(method, path, body=None, extra=None, timeout=30):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {
        "apikey": _service_key(),
        "Authorization": f"Bearer {_service_key()}",
        "content-type": "application/json",
    }
    if extra:
        headers.update(extra)
    data = json.dumps(body, default=str).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, method=method, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.status, r.read()
    except urllib.error.HTTPError as e:
        return e.code, e.read()


def download_xlsm(share_url: str) -> str:
    """Download an OneDrive/SharePoint share link's xlsm. Personal share
    URLs require &download=1 to skip the web preview AND a cookie jar to
    follow the auth redirect chain — without both we get the 55KB HTML
    sign-in page back. Mirrors the proven fetcher in cloud-sync-bid-list.py."""
    import http.cookiejar
    sep = "&" if "?" in share_url else "?"
    url = share_url + sep + "download=1"
    fd, path = tempfile.mkstemp(suffix=".xlsm")
    os.close(fd)
    print("Downloading PROJECT LIST.xlsm via SharePoint share link...")
    cookie_jar = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(
        urllib.request.HTTPCookieProcessor(cookie_jar),
        urllib.request.HTTPRedirectHandler(),
    )
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (FusionCloudCron sync-side-jobs)",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
    })
    with opener.open(req, timeout=180) as r, open(path, "wb") as f:
        total = 0
        while chunk := r.read(1 << 20):
            f.write(chunk)
            total += len(chunk)
    # Sanity: xlsx/xlsm are zip files starting with 0x504B0304 ("PK\x03\x04").
    with open(path, "rb") as f:
        magic = f.read(4)
    if magic[:2] != b"PK":
        raise SystemExit(f"Downloaded file is not an xlsm (got {magic!r}). "
                         f"Share URL likely returned HTML; check PROJECT_LIST_URL secret.")
    print(f"  downloaded {total} bytes")
    return path


def _str(v) -> str:
    return "" if v is None else str(v).strip()


def _num(v):
    if v is None or v == "": return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _date(v):
    if v is None: return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = _str(v)
    return s[:10] if s else None


def parse_tab(ws, tab_name: str, year: int, division: str) -> list[dict]:
    rows = []
    seen_jobs: set[str] = set()
    for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row, values_only=True), DATA_START_ROW):
        if not row or len(row) < 13: continue
        # Column indices (0-based):
        #   B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, J=9, M=12, O=14, S=18, Y=24
        project_name = _str(row[1] if len(row) > 1 else None)
        job_number   = _str(row[2] if len(row) > 2 else None)
        if not job_number or not project_name:
            continue
        # Skip rows whose job_number doesn't match the parent prefix (footer
        # rows, header repetitions, summary lines).
        if not re.match(rf"^{re.escape(tab_name[:-1])}\b", job_number):
            continue
        # Dedupe in case the workbook has stray duplicate rows
        if job_number in seen_jobs:
            continue
        seen_jobs.add(job_number)

        status = _str(row[12] if len(row) > 12 else None).upper()
        if status not in ACTIVE_STATUSES:
            continue   # filter at write time — only CURRENT jobs land in the table

        rows.append({
            "job_number":      job_number,
            "parent_tab":      tab_name,
            "year":            year,
            "division":        division,
            "job_suffix":      _str(row[3] if len(row) > 3 else None),
            "project_name":    project_name,
            "est_number":      _str(row[4] if len(row) > 4 else None),
            "award_date":      _date(row[5] if len(row) > 5 else None),
            "customer":        _str(row[6] if len(row) > 6 else None),
            "contact":         _str(row[7] if len(row) > 7 else None),
            "project_manager": _str(row[8] if len(row) > 8 else None),
            "foreman":         _str(row[9] if len(row) > 9 else None),
            "job_status":      status,
            "bonus_status":    _str(row[14] if len(row) > 14 else None),
            "work_description":_str(row[18] if len(row) > 18 else None),
            "contract_amount": _num(row[24] if len(row) > 24 else None),
            "payload":         {},
        })
    return rows


def replace_table(rows: list[dict]):
    """side_jobs_cloud is small (~50 rows). Wipe + reload is the simplest
    way to also drop rows whose status flipped to COMPLETE since last run.
    PostgREST DELETE requires a WHERE clause, so we use 'job_number=neq.""'
    which matches every row."""
    st, body = _sb("DELETE", f"{TABLE}?job_number=neq.")
    if st not in (200, 204):
        # neq."" doesn't quite work in some PostgREST versions; fall back
        # to id is not null using a UUID-like always-true filter.
        st, body = _sb("DELETE", f"{TABLE}?job_number=not.is.null")
        if st not in (200, 204):
            print(f"[warn] table wipe HTTP {st}: {body[:200]!r}", file=sys.stderr)

    if not rows:
        print("No active side jobs to insert.")
        return
    iso = datetime.now(timezone.utc).isoformat(timespec="seconds")
    for r in rows:
        r["generated_at"] = iso
        r["updated_at"] = iso
    st, body = _sb("POST", TABLE, body=rows,
                   extra={"Prefer": "resolution=merge-duplicates,return=minimal"})
    if st not in (200, 201, 204):
        print(f"[err] insert HTTP {st}: {body[:400]!r}", file=sys.stderr)
        sys.exit(2)


def main():
    src = os.environ.get("PROJECT_LIST_URL") or os.environ.get("PROJECT_LIST_PATH")
    if not src:
        raise SystemExit("Set PROJECT_LIST_URL or PROJECT_LIST_PATH env var.")
    path = src if os.path.exists(src) else download_xlsm(src)
    print(f"=== sync-side-jobs ===\nReading {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    all_rows: list[dict] = []
    for tab_name, (year, division) in TAB_MAP.items():
        if tab_name not in wb.sheetnames:
            print(f"  [skip] no tab {tab_name!r} in workbook")
            continue
        try:
            rows = parse_tab(wb[tab_name], tab_name, year, division)
        except Exception as e:  # noqa: BLE001
            print(f"  [warn] parse {tab_name} failed: {e}")
            continue
        print(f"  {tab_name}: {len(rows)} CURRENT side job(s)")
        all_rows.extend(rows)

    print(f"\nTotal CURRENT side jobs: {len(all_rows)}")
    replace_table(all_rows)
    print("Done.")


if __name__ == "__main__":
    main()
