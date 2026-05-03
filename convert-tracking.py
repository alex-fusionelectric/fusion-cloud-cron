"""Extract company goals from BID LIST.xlsm > TRACKING sheet.

The TRACKING sheet stores per-division annual goals. We pull the two values
the portal's existing goals model needs:

  - sent    = annual BIDDING GOAL (total bid dollars to put out)
  - awarded = annual GOAL FOR <year> (revenue target / awarded dollars)

The portal stores them as appSettings.goals[year][division] = {sent, awarded}.
Output is a JS file that fills window.__STATIC_GOALS__ + getStaticGoals().
"""
import argparse
import json
import re
import warnings
from datetime import datetime
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

SHEET_NAME = "TRACKING"

# (division, header text we expect in column A, portal field)
GOAL_FIELD_HEADERS = [
    ("sent", re.compile(r"BIDDING\s+GOAL", re.IGNORECASE)),
    ("awarded", re.compile(r"GOAL\s+FOR", re.IGNORECASE)),
    ("awardPct", re.compile(r"AWARD\s*%\s*GOAL", re.IGNORECASE)),
    ("lastYearRevenue", re.compile(r"LAST\s*YEAR.?S?\s*REV", re.IGNORECASE)),
]


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = re.sub(r"[^0-9.\-]", "", str(value))
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def find_division_blocks(ws):
    """Walk column A and pair each 'BAY' / 'SAC' marker with the next-block start."""
    blocks = []  # list of (division_label, start_row)
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1).value
        if cell is None:
            continue
        s = str(cell).strip().upper()
        if s in ("BAY", "SAC"):
            blocks.append((s, row_idx))
    return blocks


def extract_goals_for_block(ws, division, start_row, end_row):
    """Within the rows for a division block, find each known header in col A and
    pull the numeric value from col B."""
    out = {}
    for row_idx in range(start_row, end_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if not label:
            continue
        label_text = str(label).strip()
        for portal_field, pattern in GOAL_FIELD_HEADERS:
            if portal_field in out:
                continue  # don't overwrite (the bid-tracking section comes before
                          # field-tracking, so first match wins)
            if pattern.search(label_text):
                value = to_number(ws.cell(row=row_idx, column=2).value)
                if value is not None:
                    out[portal_field] = value
    return out


def detect_year(ws):
    # The sheet uses the literal "CURRENT YEAR" label and the year in row 2.
    for row_idx in range(1, min(10, ws.max_row + 1)):
        label = ws.cell(row=row_idx, column=1).value
        if label and "CURRENT YEAR" in str(label).upper():
            year_val = ws.cell(row=row_idx + 1, column=1).value
            n = to_number(year_val)
            if n:
                return int(n)
    return datetime.utcnow().year


def parse_workbook(wb, *, source_label="(unknown)"):
    """Parse a loaded BID LIST.xlsm workbook -> goals payload dict.

    Extracted so both the local pipeline (path-based) and the cloud-side
    pipeline (SharePoint share-link bytes) share one parse implementation.
    """
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")
    ws = wb[SHEET_NAME]

    year = detect_year(ws)
    blocks = find_division_blocks(ws)
    if not blocks:
        raise SystemExit("No BAY/SAC division blocks found in TRACKING sheet")

    seen_divs = set()
    goals_by_division = {}
    for i, (division, start_row) in enumerate(blocks):
        if division in seen_divs:
            continue
        seen_divs.add(division)
        end_row = blocks[i + 1][1] - 1 if i + 1 < len(blocks) else ws.max_row
        goals = extract_goals_for_block(ws, division, start_row, end_row)
        if goals:
            goals_by_division[division] = goals

    return {
        "generatedAt": datetime.utcnow().isoformat() + "Z",
        "source": source_label,
        "year": year,
        "goals": { str(year): goals_by_division },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.excel, data_only=True)
    payload = parse_workbook(wb, source_label=str(args.excel))
    year = payload.get("year")
    goals_by_division = payload.get("goals", {}).get(str(year), {})
    payload_json = json.dumps(payload, indent=2)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    js = (
        "// AUTO-GENERATED -- do not edit. Regenerate with fusion-portal/Convert-Tracking.ps1\n"
        f"window.__STATIC_GOALS__ = {payload_json};\n"
        "window.getStaticGoals = function(){\n"
        "  return window.__STATIC_GOALS__ || null;\n"
        "};\n"
    )
    out_path.write_text(js, encoding="utf-8")
    json_path = out_path.with_suffix(".json")
    json_path.write_text(payload_json, encoding="utf-8")
    print(
        f"Wrote goals for year {year}, divisions {list(goals_by_division.keys())} "
        f"to {out_path} and {json_path}"
    )


if __name__ == "__main__":
    main()
