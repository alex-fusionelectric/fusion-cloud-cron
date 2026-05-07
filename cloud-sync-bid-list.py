"""Cloud-side BID LIST parser: downloads BID LIST.xlsm from a SharePoint
share URL (anonymous &download=1) and produces a bids-data.json file
identical in shape to what sync-bid-list.py emits locally.

Purpose: prove parse equivalence against the live local pipeline. Once
this script's output diffs cleanly against the canonical
fusion-pm panel/src/assets/bids-data.json, we know the parse logic is
portable and we can wire it into a cloud cron (GitHub Actions, etc.) to
retire the local PC dependency entirely.

This file deliberately does NOT write to Supabase or anywhere else --
it's pure read-side validation. Step 3 (Supabase write + cron host)
comes after the diff is green.

USAGE
  python cloud-sync-bid-list.py \
      --url "https://fusionelectricinc-my.sharepoint.com/:x:/g/personal/.../FILE?e=ABC" \
      --out cloud-bids-data.json

The --url is a "Anyone with the link can view" share URL from OneDrive /
SharePoint. The script appends &download=1 automatically. The local
sync-bid-list.py module is imported wholesale -- we don't duplicate any
parse logic, only how the workbook bytes are obtained.
"""

import argparse
import http.cookiejar
import io
import json
import os
import re
import sys
import urllib.error
import urllib.request
import urllib.parse
from datetime import datetime
from pathlib import Path

# Reuse the canonical parser. We import everything we need by name so
# behavior stays in lockstep with the local pipeline -- if Dave adds a
# new sheet column tomorrow and sync-bid-list.py is updated to read it,
# this script picks the change up automatically with no edits.
SCRIPTS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPTS_DIR))

import pandas as pd  # noqa: E402
import openpyxl  # noqa: E402

import sync_bid_list_compat  # noqa: E402  -- shim, see bottom of file


def fetch_xlsm_bytes(share_url):
    """Download the xlsm via the share link's direct-download endpoint.
    OneDrive/SharePoint personal share links accept &download=1 to bypass
    the web preview and serve the file. Returns raw bytes; caller wraps
    in BytesIO for openpyxl/pandas.

    SharePoint's redirect chain sets an auth cookie on hop 1 and requires
    it on hop 2 -- without a cookie jar urllib gets back the 55 KB
    HTML preview/sign-in page instead of the file. Use an opener with
    HTTPCookieProcessor to follow the chain correctly. Sanity-check the
    response with the xlsx/zip magic bytes so a config regression
    (revoked share, stricter tenant policy, expired URL) fails loudly
    instead of dumping an unparseable HTML blob into pandas.
    """
    sep = "&" if "?" in share_url else "?"
    url = share_url + sep + "download=1"
    # URL deliberately scrubbed -- this script runs in a public GitHub
    # repo's workflow logs; even a truncated SharePoint URL leaks the
    # tenant + file owner. Useful info kept (bytes received) is logged
    # after the download completes.
    print("Downloading SharePoint xlsm...")

    cookie_jar = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(
        urllib.request.HTTPCookieProcessor(cookie_jar),
        urllib.request.HTTPRedirectHandler(),
    )
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36",
        "Accept": "*/*",
    })
    with opener.open(req, timeout=60) as resp:
        if resp.status != 200:
            raise SystemExit(f"download failed: HTTP {resp.status}")
        data = resp.read()

    if not data.startswith(b"PK"):
        # xlsx/xlsm files are zip archives -- magic is "PK\x03\x04". HTML
        # starts with "<", a sign-in page is usually < 200 KB. Surface
        # both clues so the operator can fix the share link fast.
        head = data[:120].decode("utf-8", errors="replace")
        raise SystemExit(
            f"download returned non-xlsm content ({len(data):,} bytes, "
            f"starts with {head!r}). Likely the share link expired, was "
            f"revoked, or your tenant blocked anonymous access."
        )

    print(f"  got {len(data):,} bytes ({len(data)/1024/1024:.2f} MB)")
    return data


def parse_workbook(xlsm_bytes):
    """Run the exact same parse pipeline as sync-bid-list.py main(),
    but reading from bytes instead of a path. Returns the {generatedAt,
    source, estimatorDivisions, bids} dict ready to be dumped to JSON."""
    # pandas + openpyxl both accept BytesIO. Each needs its own buffer
    # since they read the stream destructively.
    bio_pd_bids = io.BytesIO(xlsm_bytes)
    bio_pd_follow = io.BytesIO(xlsm_bytes)
    bio_pd_arch = io.BytesIO(xlsm_bytes)
    bio_xl = io.BytesIO(xlsm_bytes)

    bids_df   = pd.read_excel(bio_pd_bids,   sheet_name="BIDS",       header=7, engine="openpyxl")
    follow_df = pd.read_excel(bio_pd_follow, sheet_name="FOLLOW UPS", header=3, engine="openpyxl")
    archive_df= pd.read_excel(bio_pd_arch,   sheet_name="ARCHIVE",    header=1, engine="openpyxl")

    # Hyperlink scan -- openpyxl preserves cell.hyperlink which pandas drops.
    documents_url_by_est = _hyperlink_map_from_bytes(bio_xl)
    print(f"Found {len(documents_url_by_est)} bid name hyperlinks (OneDrive / Dropbox / plan-room).")

    excel_estimators = sync_bid_list_compat.extract_estimators(bids_df, follow_df, archive_df)
    # Skip the HTML CORE_ESTIMATORS update -- that's a side effect of the
    # local pipeline that has nothing to do with parse correctness, and
    # the cloud script doesn't have the index.html file path to update.

    allowed = excel_estimators or None
    pe_map = sync_bid_list_compat.build_pe_map(bids_df, follow_df, archive_df)
    division_map = sync_bid_list_compat.build_estimator_division_map([(archive_df, 2), (follow_df, 1)])

    bids = []

    def apply_division_map(rows):
        for row in rows:
            if not row.get("division") and row.get("estimator"):
                est = str(row.get("estimator", "")).strip().upper()
                row["division"] = division_map.get(est, "") or sync_bid_list_compat.division_from_estimator(est)
            est_num = str(row.get("estNumber", "")).strip().upper()
            url = documents_url_by_est.get(est_num)
            if url:
                row["documentsUrl"] = url
        return rows

    bids.extend(apply_division_map(sync_bid_list_compat.parse_bids(bids_df, allowed, pe_map)))
    bids.extend(apply_division_map(sync_bid_list_compat.parse_follow(follow_df, allowed, pe_map)))
    bids.extend(apply_division_map(sync_bid_list_compat.parse_archive(archive_df, allowed, pe_map)))

    # Same dedup + id stamping as sync-bid-list.py:457-468.
    seen = set()
    deduped = []
    for b in bids:
        key = (str(b.get("estNumber", "")).upper(), b.get("projectName", ""), b.get("status", ""))
        if key in seen:
            continue
        seen.add(key)
        b["id"] = f"{key[0]}::{b.get('status','')}::{len(deduped)}"
        b["createdAt"] = datetime.utcnow().isoformat() + "Z"
        b["updatedAt"] = datetime.utcnow().isoformat() + "Z"
        deduped.append(b)

    return {
        "generatedAt": datetime.utcnow().isoformat() + "Z",
        "source": "cloud:sharepoint-share-url",
        "estimatorDivisions": division_map,
        "bids": deduped,
    }


# Matches =HYPERLINK("url", "display") or =HYPERLINK("url"). Tolerates
# leading/trailing whitespace and case differences. Captures the URL only.
_HYPERLINK_FORMULA_RX = re.compile(r'^\s*=\s*HYPERLINK\s*\(\s*"([^"]+)"', re.IGNORECASE)


def _hyperlink_map_from_bytes(xlsm_bytes_io):
    """In-memory equivalent of sync-bid-list.py:_build_documents_url_map.
    Same sheet/column/row offsets so the output keys + URLs match.

    Two-pass scan per row:
      1. Native Hyperlink object on the project-name cell (the standard
         "right-click > Edit Hyperlink" path that Excel uses by default).
      2. Fallback: =HYPERLINK("url","text") formula in the cell. Some
         users (and our xlwings tooling when COM Hyperlinks.Add 500s on
         OneDrive-hosted workbooks) write the link this way instead.
         Without this fallback, the cron would see "no hyperlink" on
         such cells and wipe documents_url to NULL on every sync.

    For the formula path we need a separate workbook load with
    data_only=False so cell.value returns the formula text instead of
    the calculated display string.
    """
    out = {}
    sheet_configs = [
        ("BIDS",       8, "B", "C"),
        ("FOLLOW UPS", 4, "B", "C"),
        ("ARCHIVE",    2, "B", "C"),
    ]
    try:
        wb = openpyxl.load_workbook(xlsm_bytes_io, data_only=True)
    except Exception as exc:  # noqa: BLE001
        print(f"[warn] could not open workbook for hyperlink scan: {exc}")
        return out

    # Second pass needs the same bytes but in formula-mode. Rewind the
    # BytesIO so the second load sees the start of the file.
    formula_wb = None
    try:
        if hasattr(xlsm_bytes_io, "seek"):
            xlsm_bytes_io.seek(0)
        formula_wb = openpyxl.load_workbook(xlsm_bytes_io, data_only=False)
    except Exception as exc:  # noqa: BLE001
        print(f"[warn] could not open workbook for HYPERLINK-formula fallback: {exc}")

    for sheet_name, data_start, name_col, est_col in sheet_configs:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws_formula = formula_wb[sheet_name] if (formula_wb is not None and sheet_name in formula_wb.sheetnames) else None
        for row in range(data_start, ws.max_row + 1):
            name_cell = ws[f"{name_col}{row}"]
            est_cell  = ws[f"{est_col}{row}"]
            est = str(est_cell.value or "").strip().upper()
            if not est:
                continue
            # Pass 1: native Hyperlink object.
            url = None
            if name_cell.hyperlink and name_cell.hyperlink.target:
                url = name_cell.hyperlink.target
            # Pass 2: HYPERLINK() formula fallback.
            elif ws_formula is not None:
                raw = ws_formula[f"{name_col}{row}"].value
                if isinstance(raw, str):
                    m = _HYPERLINK_FORMULA_RX.match(raw)
                    if m:
                        url = m.group(1)
            if url and est not in out:
                out[est] = url
    wb.close()
    if formula_wb is not None:
        formula_wb.close()
    return out


SUPABASE_URL = "https://dltuvsdwrujjsmiotaxy.supabase.co"
SUPABASE_TABLE = "bids_cloud"


def _resolve_service_key():
    """Pull the Supabase service-role key from the standard places. Order:
      1. SUPABASE_SERVICE_KEY env var (the cloud cron path -- GitHub Actions
         secret gets injected as an env var).
      2. fusion-bid-list/supabase-service-key.txt (the local-PC path -- same
         file Push-ToSupabase.ps1 already uses, so no new credentials to manage).
    Returns the key string or raises SystemExit on miss."""
    key = os.environ.get("SUPABASE_SERVICE_KEY", "").strip()
    if key:
        return key
    # Repo-relative fallback. SCRIPTS_DIR is fusion-pm panel/scripts/, so
    # ../../fusion-bid-list/supabase-service-key.txt.
    fallback = SCRIPTS_DIR.parent.parent / "fusion-bid-list" / "supabase-service-key.txt"
    if fallback.is_file():
        contents = fallback.read_text(encoding="utf-8").strip()
        if contents:
            return contents
    raise SystemExit(
        "no Supabase service key found. Set SUPABASE_SERVICE_KEY env var "
        f"or create {fallback}."
    )


def _bid_to_row(bid, generated_at):
    """Project a bid dict (bids-data.json shape) into a bids_cloud row.
    Native columns mirror what the portal queries against; the full bid
    object lives in `payload` so future schema additions don't require
    a migration."""
    return {
        "id":               bid.get("id"),
        "est_number":       bid.get("estNumber") or None,
        "estimator":        bid.get("estimator") or None,
        "project_engineer": bid.get("projectEngineer") or None,
        "project_name":     bid.get("projectName") or None,
        "client_gc":        bid.get("clientGc") or None,
        "bid_amount":       bid.get("bidAmount"),
        "awarded_amount":   bid.get("awardedAmount"),
        "bid_due_date":     bid.get("bidDueDate"),
        "status":           bid.get("status") or None,
        "outcome":          bid.get("outcome") or None,
        "date_sent":        bid.get("dateSent"),
        "follow_up_date":   bid.get("followUpDate"),
        "date_awarded":     bid.get("dateAwarded"),
        "division":         bid.get("division") or None,
        "documents_url":    bid.get("documentsUrl") or None,
        "payload":          bid,
        "generated_at":     generated_at,
        "updated_at":       generated_at,
    }


def write_to_supabase(output, *, batch_size=200):
    """Upsert-then-purge pattern: ON CONFLICT (id) DO UPDATE so the table
    is never empty mid-run, then delete any row whose updated_at predates
    this run (meaning it was removed from the source workbook).
    """
    key = _resolve_service_key()
    api = f"{SUPABASE_URL}/rest/v1/{SUPABASE_TABLE}"
    headers_upsert = {
        "apikey":        key,
        "Authorization": f"Bearer {key}",
        "Content-Type":  "application/json",
        "Prefer":        "resolution=merge-duplicates,return=minimal",
    }
    headers_delete = {
        "apikey":        key,
        "Authorization": f"Bearer {key}",
        "Content-Type":  "application/json",
        "Prefer":        "return=minimal",
    }

    bids = output.get("bids", [])
    generated_at = output.get("generatedAt")
    rows = [_bid_to_row(b, generated_at) for b in bids]

    # 1. Upsert all rows. Every row carries updated_at = generated_at so
    # the stale-purge step below can identify rows not in this run.
    print(f"Upserting {len(rows)} row(s) in batches of {batch_size}...")
    sent = 0
    for i in range(0, len(rows), batch_size):
        chunk = rows[i:i + batch_size]
        body = json.dumps(chunk).encode("utf-8")
        req = urllib.request.Request(api, method="POST", headers=headers_upsert, data=body)
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                if resp.status not in (200, 201, 204):
                    raise SystemExit(f"POST failed mid-batch: HTTP {resp.status}")
        except urllib.error.HTTPError as e:
            err = e.read().decode("utf-8", errors="replace")
            raise SystemExit(f"POST failed at row {i}: HTTP {e.code} {err}")
        sent += len(chunk)
        print(f"  ... {sent}/{len(rows)}")
    print(f"Wrote {sent} row(s) to public.{SUPABASE_TABLE}.")

    # 2. Purge stale rows (bids removed from the workbook since last run).
    # Rows from this run have updated_at == generated_at; anything older
    # was not present in the latest parse.
    stale_url = api + "?updated_at=neq." + urllib.parse.quote(generated_at, safe="")
    req = urllib.request.Request(stale_url, method="DELETE", headers=headers_delete)
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            if resp.status not in (200, 204):
                raise SystemExit(f"DELETE stale failed: HTTP {resp.status}")
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise SystemExit(f"DELETE stale failed: HTTP {e.code} {body}")
    print(f"Purged stale rows (updated_at != {generated_at}).")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--url", required=True, help="OneDrive/SharePoint share link (Anyone-with-link).")
    ap.add_argument("--out", default=None,
                    help="Optional: write a JSON snapshot of the parse output. "
                         "Useful for diff-bids.py validation; not needed in production.")
    ap.add_argument("--supabase-write", action="store_true",
                    help="Truncate + insert all rows into public.bids_cloud. "
                         "Reads service key from SUPABASE_SERVICE_KEY env var "
                         "or fusion-bid-list/supabase-service-key.txt.")
    args = ap.parse_args()

    xlsm = fetch_xlsm_bytes(args.url)
    output = parse_workbook(xlsm)

    if args.out:
        out_path = Path(args.out)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(output, indent=2), encoding="utf-8")
        print(f"Wrote {len(output['bids'])} bids to {out_path}")

    if args.supabase_write:
        write_to_supabase(output)

    if not args.out and not args.supabase_write:
        print("[warn] no output flag set (--out or --supabase-write). "
              "Parse succeeded but result was discarded.", file=sys.stderr)


if __name__ == "__main__":
    main()
